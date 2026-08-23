"""NM-V1 rule engine application + blinded 120-case expert-audit package.

Applies the frozen structural decision engine to all 1,283 blocks using REFERENCE evidence
only. The HMM annotation is read solely to define the S_DISC audit stratum and is never used
to assign a label, never written into the audit package, and never shown to the adjudicator.
"""
import argparse, collections, csv, datetime, hashlib, html, json, os, sys

VERSION = "nmv1_apply_rule_engine_v1.0.0"
ENGINE_SHA = "ed5db383bb0afe1a1a8433886d6666fe72c324975de99c6763a37824d51c2bee"
OUTCOMES = ["chromosomal_mobile_supported", "chromosomal_quiescent_supported",
            "integron_associated_supported", "IS_associated_supported",
            "multiple_MGE_evidence_supported", "neither_classification_supported",
            "biologically_indeterminate"]
MOBILE_RULES = {"A", "B", "C", "D"}


def sha256_file(p):
    h = hashlib.sha256()
    with open(p, "rb") as fh:
        for c in iter(lambda: fh.read(1 << 20), b""):
            h.update(c)
    return h.hexdigest()


def dsort(items, seed, salt):
    return sorted(items, key=lambda x: hashlib.sha256(
        ("%s|%d|%s" % (x, seed, salt)).encode()).hexdigest())


def classify(p):
    """Frozen rule order. First match wins. Returns (rule_id, label, reason)."""
    if p["tool_problem"]:
        return "F1", "biologically_indeterminate", "missing or failed tool output"
    if p["boundary_problem"]:
        return "F2", "biologically_indeterminate", "truncated or origin-wrapped block"
    if p["IS_strong"] and p["INT_strong"]:
        return "A", "multiple_MGE_evidence_supported", "complete IS with bilateral TIRs and complete integron with intI+attC"
    if p["IS_strong"] and not p["INT_strong"]:
        return "B", "IS_associated_supported", "complete IS, complete transposase ORF, resolved bilateral TIRs"
    if p["INT_strong"] and not p["IS_strong"]:
        return "C", "integron_associated_supported", "complete integron with integrase and attC"
    if (p["IS_complete_n"] > 0 or p["INT_complete"]) and not (p["IS_strong"] or p["INT_strong"]):
        return "D", "chromosomal_mobile_supported", "complete-level mobile evidence not exclusively assignable"
    if p["IS_partial_only"] or p["INT_incomplete_only"]:
        return "F3", "biologically_indeterminate", "partial-only element or incomplete integron"
    if p["no_reference_evidence"] and not p["tool_problem"] and not p["boundary_problem"]:
        return "E", "chromosomal_quiescent_supported", "no structural MGE evidence, both tools completed, evaluable"
    return "F4", "biologically_indeterminate", "conflicting evidence unresolved by the frozen rules"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo", required=True)
    ap.add_argument("--engine", required=True)
    ap.add_argument("--evidence", required=True)
    ap.add_argument("--results", required=True)
    ap.add_argument("--outdir", required=True)
    a = ap.parse_args()
    if sha256_file(a.engine) != ENGINE_SHA:
        print("REFUSING: rule engine digest mismatch"); sys.exit(1)
    E = json.load(open(a.engine, encoding="utf-8"))
    if not E.get("frozen_before_application"):
        print("REFUSING: engine does not assert pre-application freeze"); sys.exit(1)
    SEED = E["expert_audit"]["seed"]
    print("%s | engine %s verified" % (VERSION, ENGINE_SHA[:16]))

    ev = {r["block_id"]: r for r in csv.DictReader(open(a.evidence, encoding="utf-8"),
                                                   delimiter="\t")}
    O = os.path.join(a.repo, "audit/data/derived/pr_context/out")
    blk = {r["block_id"]: r for r in csv.DictReader(
        open(os.path.join(O, "shared_context_blocks.tsv"), encoding="utf-8"), delimiter="\t")}

    ise = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(a.results, "isescan_hits.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        ise[r["block_id"]].append(r)
    inf = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(a.results, "integronfinder_features.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        inf[r["block_id"]].append(r)

    def iv(x):
        try:
            return int(float(x))
        except (TypeError, ValueError):
            return 0

    gt = []
    for b, e in ev.items():
        hits = ise.get(b, [])
        comp = [h for h in hits if h.get("type") == "c"]
        bil = [h for h in comp if iv(h.get("start1")) and iv(h.get("end1"))
               and iv(h.get("start2")) and iv(h.get("end2")) and iv(h.get("irLen")) > 0]
        orf = [h for h in comp if iv(h.get("orfLen")) > 0]
        fe = inf.get(b, [])
        cf = [f for f in fe if f.get("type") == "complete"]
        intI = any(f.get("annotation") == "intI" for f in cf)
        attC = any(f.get("type_elt") == "attC" for f in cf)
        p = {"IS_complete_n": len(comp), "IS_bilateral_TIR_n": len(bil),
             "IS_complete_orf_n": len(orf),
             "IS_strong": len(bil) > 0 and len(orf) > 0,
             "IS_partial_only": iv(e["isescan_partial"]) > 0 and len(comp) == 0,
             "INT_complete": iv(e["if_complete"]) > 0,
             "INT_intI": intI, "INT_attC": attC,
             "INT_strong": iv(e["if_complete"]) > 0 and intI and attC,
             "INT_incomplete_only": (iv(e["if_calin"]) + iv(e["if_in0"])) > 0
                                    and iv(e["if_complete"]) == 0,
             "no_reference_evidence": iv(e["isescan_n"]) == 0
                                      and (iv(e["if_complete"]) + iv(e["if_calin"])
                                           + iv(e["if_in0"])) == 0,
             "boundary_problem": e["truncated"] == "yes" or e["wrapped"] == "yes",
             "tool_problem": e["tool_status"] != "ok"}
        rid, lab, why = classify(p)
        gt.append({"block_id": b, "rule_id": rid, "rule_based_label": lab, "rule_reason": why,
                   "evaluable_status": "indeterminate" if lab == "biologically_indeterminate"
                                       else "evaluable",
                   "IS_complete_n": p["IS_complete_n"],
                   "IS_bilateral_TIR_n": p["IS_bilateral_TIR_n"],
                   "IS_complete_orf_n": p["IS_complete_orf_n"],
                   "IS_partial_n": iv(e["isescan_partial"]),
                   "INT_complete_n": iv(e["if_complete"]), "INT_calin_n": iv(e["if_calin"]),
                   "INT_in0_n": iv(e["if_in0"]),
                   "INT_intI": "yes" if intI else "no", "INT_attC": "yes" if attC else "no",
                   "topology": e["topology"], "truncated": e["truncated"],
                   "wrapped": e["wrapped"], "tool_status": e["tool_status"],
                   "hmm_is": iv(e["hmm_is"]), "hmm_integron": iv(e["hmm_integron"]),
                   "stratum": e["stratum"], "species": e["species"],
                   "bioproject": e["bioproject"]})
    gt.sort(key=lambda r: r["block_id"])
    cnt = collections.Counter(r["rule_id"] for r in gt)
    lab = collections.Counter(r["rule_based_label"] for r in gt)
    print("\n=== RULE-BASED GROUND TRUTH, all %d blocks ===" % len(gt))
    for k in ("A", "B", "C", "D", "E", "F1", "F2", "F3", "F4"):
        r = [x for x in E["rule_order"] if x["id"] == k][0]
        print("  %-3s %-38s %5d" % (k, r["label"], cnt.get(k, 0)))
    print("  %-42s %5d" % ("TOTAL", sum(cnt.values())))
    print("\n  by label:")
    for k, v in lab.most_common():
        print("     %-38s %5d" % (k, v))
    print("  evaluable %d | indeterminate %d"
          % (sum(1 for r in gt if r["evaluable_status"] == "evaluable"),
             sum(1 for r in gt if r["evaluable_status"] == "indeterminate")))

    os.makedirs(a.outdir, exist_ok=True)
    gp = os.path.join(a.outdir, "NMV1_RULE_BASED_GROUND_TRUTH.tsv")
    gcols = ["block_id", "rule_id", "rule_based_label", "evaluable_status", "rule_reason",
             "IS_complete_n", "IS_bilateral_TIR_n", "IS_complete_orf_n", "IS_partial_n",
             "INT_complete_n", "INT_calin_n", "INT_in0_n", "INT_intI", "INT_attC",
             "topology", "truncated", "wrapped", "tool_status"]
    with open(gp, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\t".join(gcols) + "\n")
        for r in gt:
            fh.write("\t".join(str(r[c]) for c in gcols) + "\n")
    print("\n  ground truth table: %s" % sha256_file(gp))

    # ================= audit strata =================
    idx = {r["block_id"]: r for r in gt}
    pools = {
      "S_IS": [r["block_id"] for r in gt if r["rule_id"] == "B"],
      "S_INT": [r["block_id"] for r in gt if r["rule_id"] == "C"],
      "S_MULTI": [r["block_id"] for r in gt if r["rule_id"] == "A"],
      "S_QUIET": [r["block_id"] for r in gt if r["rule_id"] == "E"],
      "S_DISC": [r["block_id"] for r in gt
                 if r["rule_id"] in MOBILE_RULES | {"E"}
                 and ((r["hmm_is"] + r["hmm_integron"]) > 0) != (r["rule_id"] in MOBILE_RULES)],
      "S_INDET": [r["block_id"] for r in gt if r["rule_id"].startswith("F")],
    }
    order = [s["id"] for s in E["expert_audit"]["strata"]]
    target = {s["id"]: s["target"] for s in E["expert_audit"]["strata"]}
    print("\n=== AUDIT STRATA (pools recorded before selection) ===")
    for s in order:
        print("  %-9s pool %5d  target %3d" % (s, len(pools[s]), target[s]))

    # shortfall + proportional reallocation, before any human decision
    short = {s: max(0, target[s] - len(pools[s])) for s in order}
    deficit = sum(short.values())
    alloc = {s: min(target[s], len(pools[s])) for s in order}
    if deficit:
        recv = [s for s in order if len(pools[s]) > alloc[s]]
        cap = {s: len(pools[s]) - alloc[s] for s in recv}
        tot = sum(cap.values())
        print("  shortfall %d -> reallocating proportionally across %d strata"
              % (deficit, len(recv)))
        given = 0
        for i, s in enumerate(recv):
            add = deficit - given if i == len(recv) - 1 else int(round(deficit * cap[s] / tot))
            add = min(add, cap[s], deficit - given)
            alloc[s] += add; given += add
        if given < deficit:
            for s in recv:
                while given < deficit and alloc[s] < len(pools[s]):
                    alloc[s] += 1; given += 1
    sel = {}; used = set()
    for s in order:
        avail = [b for b in pools[s] if b not in used]
        pick = dsort(avail, SEED, s)[:alloc[s]]
        sel[s] = sorted(pick); used.update(pick)
        print("  %-9s allocated %3d  selected %3d" % (s, alloc[s], len(pick)))
    audit = sorted(used)
    print("  AUDIT TOTAL: %d (cap %d) | mutually exclusive: %s"
          % (len(audit), E["expert_audit"]["max_cases"],
             len(audit) == sum(len(v) for v in sel.values())))

    # ---- audit manifest, before the package ----
    amp = os.path.join(a.outdir, "NMV1_AUDIT_MANIFEST.json")
    AM = {"manifest": "NMV1_AUDIT_MANIFEST", "builder": VERSION, "seed": SEED,
          "frozen_utc": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
          "engine_sha256": ENGINE_SHA,
          "selection_rule": E["expert_audit"]["selection_rule"],
          "strata": {s: {"pool": len(pools[s]), "target": target[s], "allocated": alloc[s],
                         "selected": len(sel[s]),
                         "token_hash": hashlib.sha256("|".join(sorted(sel[s])).encode()).hexdigest()}
                     for s in order},
          "total_selected": len(audit),
          "all_token_hash": hashlib.sha256("|".join(audit).encode()).hexdigest(),
          "hmm_used_for_selection_only": True,
          "hmm_never_shown_to_adjudicator": True}
    json.dump(AM, open(amp, "w", encoding="utf-8", newline="\n"), indent=2)
    print("  audit manifest: %s" % sha256_file(amp))

    # ================= blinded package =================
    argb = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(O, "arg_mge_neighbourhood.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        rep = r["replicon_accession"]
        for b in audit:
            s = blk[b]
            if s["replicon_accession"] != rep:
                continue
            bs, be = int(s["block_start"]), int(s["block_end"])
            gs, ge = int(r["gene_start"]), int(r["gene_end"])
            if bs <= gs and be >= ge:
                argb[b].append((gs - bs + 1, ge - bs + 1, r["strand"]))
                break
    hmmb = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(O, "mge_feature_inventory.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        b = r["block_id"]
        if b not in set(audit):
            continue
        bs = int(blk[b]["block_start"])
        cl = r["feature_class"].lower()
        hmmb[b].append({"kind": "transposase" if ("is" in cl or "transpos" in cl) else "integrase",
                        "name": r["feature_name"], "beg": int(r["chrom_start"]) - bs + 1,
                        "end": int(r["chrom_end"]) - bs + 1, "strand": r["strand"],
                        "evalue": r["evalue"]})

    tok = {b: "NMV1A-%03d" % (i + 1) for i, b in enumerate(dsort(audit, SEED, "audittoken"))}

    def feats_for(b):
        L = int(blk[b]["block_span_bp"])
        av = argb.get(b, [])
        def dist(x, y):
            if not av:
                return ""
            return min(0 if (x <= ae and y >= ab) else min(abs(x - ae), abs(ab - y))
                       for ab, ae, _ in av)
        F = []
        for f in hmmb.get(b, []):
            F.append({"m": "X", "t": f["kind"], "l": f["name"], "b": f["beg"], "e": f["end"],
                      "s": f["strand"], "v": f["evalue"], "x": "",
                      "d": dist(f["beg"], f["end"])})
        for h in ise.get(b, []):
            ir = ("L:%s-%s R:%s-%s len=%s id=%s" % (h["start1"], h["end1"], h["start2"],
                                                    h["end2"], h["irLen"], h["irId"])
                  if iv(h.get("irLen")) > 0 else "none")
            F.append({"m": "Y", "t": "IS_complete" if h.get("type") == "c" else "IS_partial",
                      "l": "%s/%s" % (h.get("family", ""), h.get("cluster", "")),
                      "b": iv(h["isBegin"]), "e": iv(h["isEnd"]), "s": h.get("strand", ""),
                      "v": h.get("E_value", ""),
                      "x": "TIR %s | ORF %s-%s len %s" % (ir, h.get("orfBegin", ""),
                                                          h.get("orfEnd", ""), h.get("orfLen", "")),
                      "d": dist(iv(h["isBegin"]), iv(h["isEnd"]))})
        for f in inf.get(b, []):
            pb, pe = iv(f["pos_beg"]), iv(f["pos_end"])
            if not pb and not pe:
                continue
            t = "attC" if f.get("type_elt") == "attC" else (
                "integrase" if f.get("annotation") == "intI" else "cassette_protein")
            F.append({"m": "Z", "t": t, "l": "%s [%s]" % (f.get("id_integron", ""), f.get("type", "")),
                      "b": pb, "e": pe, "s": f.get("strand", ""), "v": f.get("evalue", ""),
                      "x": "model=%s" % f.get("model", ""), "d": dist(pb, pe)})
        F.sort(key=lambda z: z["b"])
        return L, av, F

    def svg(L, av, F):
        W, LH = 980, 22
        tr = ["ARG", "X", "Y", "Z"]
        H = 60 + len(tr) * LH
        sc = lambda x: 40 + (W - 80) * max(0, min(L, x)) / max(L, 1)
        p = ['<svg xmlns="http://www.w3.org/2000/svg" width="%d" height="%d" '
             'font-family="ui-monospace,monospace" font-size="11">' % (W, H),
             '<rect width="%d" height="%d" fill="#fff"/>' % (W, H),
             '<line x1="%.1f" y1="28" x2="%.1f" y2="28" stroke="#333"/>' % (sc(0), sc(L))]
        for k in range(6):
            x = sc(L * k / 5.0)
            p.append('<line x1="%.1f" y1="24" x2="%.1f" y2="32" stroke="#333"/>' % (x, x))
            p.append('<text x="%.1f" y="20" text-anchor="middle" fill="#333">%d</text>'
                     % (x, int(L * k / 5.0)))
        p.append('<text x="8" y="20" fill="#333">bp</text>')
        yt = {t: 44 + i * LH for i, t in enumerate(tr)}
        for t in tr:
            p.append('<text x="8" y="%d" fill="#555">%s</text>' % (yt[t] + 10, t))
            p.append('<line x1="%.1f" y1="%d" x2="%.1f" y2="%d" stroke="#eee"/>'
                     % (sc(0), yt[t] + 7, sc(L), yt[t] + 7))
        for ab, ae, _ in av:
            p.append('<rect x="%.1f" y="%d" width="%.1f" height="12" fill="#9e9e9e" stroke="#555"/>'
                     % (sc(ab), yt["ARG"], max(2, sc(ae) - sc(ab))))
        col = {"transposase": "#7fb3d5", "integrase": "#c39bd3", "IS_complete": "#5d8aa8",
               "IS_partial": "#a9cce3", "attC": "#f0b27a", "cassette_protein": "#d5dbdb"}
        for f in F:
            y = yt.get(f["m"], yt["X"]); x1, x2 = sc(f["b"]), sc(f["e"])
            p.append('<rect x="%.1f" y="%d" width="%.1f" height="12" fill="%s" stroke="#444">'
                     '<title>%s %s-%s</title></rect>'
                     % (x1, y, max(2, x2 - x1), col.get(f["t"], "#ccc"),
                        html.escape(f["t"]), f["b"], f["e"]))
            if f["t"] == "IS_complete" and "TIR L:" in f["x"]:
                for tx in (x1, x2):
                    p.append('<rect x="%.1f" y="%d" width="4" height="12" fill="#1b2631"/>' % (tx - 2, y))
        p.append('</svg>')
        return "".join(p)

    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "audit"
    hdr = ["token", "adjudication", "reason", "topology", "boundary_warning", "window_length_bp",
           "arg_intervals", "methodX_transposase", "methodX_integrase", "methodY_IS_complete",
           "methodY_IS_partial", "methodY_bilateral_TIR", "methodY_TIR_detail", "methodY_ORFs",
           "methodZ_integrase", "methodZ_attC", "methodZ_structure",
           "nearest_feature_distance_bp", "tool_status", "feature_map"]
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = Alignment(wrap_text=True, vertical="top")
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(OUTCOMES), allow_blank=False)
    dv.prompt = "Exactly one outcome. biologically_indeterminate is legitimate."
    ws.add_data_validation(dv)
    cases = []
    for b in dsort(audit, SEED, "audittoken"):
        L, av, F = feats_for(b)
        g = idx[b]; t = tok[b]
        bil = sum(1 for f in F if f["t"] == "IS_complete" and "TIR L:" in f["x"])
        tir = "; ".join(f["x"].split(" | ")[0][4:] for f in F
                        if f["t"].startswith("IS") and "TIR L:" in f["x"])[:250]
        orfs = "; ".join(f["x"].split(" | ")[1] for f in F
                         if f["t"].startswith("IS") and " | " in f["x"])[:250]
        struct = ",".join(sorted({f["l"].split("[")[-1].rstrip("]") for f in F if f["m"] == "Z"})) or "none"
        ds = [f["d"] for f in F if f["d"] != ""]
        ws.append([t, "", "", g["topology"],
                   "yes" if (g["truncated"] == "yes" or g["wrapped"] == "yes") else "no", L,
                   "; ".join("%d-%d" % (x[0], x[1]) for x in av[:6]) or "none",
                   sum(1 for f in F if f["m"] == "X" and f["t"] == "transposase"),
                   sum(1 for f in F if f["m"] == "X" and f["t"] == "integrase"),
                   sum(1 for f in F if f["t"] == "IS_complete"),
                   sum(1 for f in F if f["t"] == "IS_partial"), bil,
                   tir or "none", orfs or "none",
                   sum(1 for f in F if f["m"] == "Z" and f["t"] == "integrase"),
                   sum(1 for f in F if f["t"] == "attC"), struct,
                   min(ds) if ds else "", g["tool_status"], "casebook: %s" % t])
        r = ws.max_row
        dv.add(ws.cell(row=r, column=2))
        ws.cell(row=r, column=20).hyperlink = "NMV1_AUDIT_CASEBOOK.html#%s" % t
        ws.cell(row=r, column=20).font = Font(color="0000EE", underline="single")
        cases.append((t, b, L, av, F, g))
    for col, w in zip("ABCDEFGHIJKLMNOPQRST",
                      [12, 34, 44, 10, 12, 14, 24, 13, 13, 13, 13, 13, 32, 32, 12, 10, 16, 14, 14, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B2"
    ws2 = wb.create_sheet("rubric")
    ws2.append(["NM-V1 expert-audit rubric"]); ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])
    ws2.append(["Judge only the structural evidence shown. Species, study, gene identity, the "
                "original classification and any machine-derived label are withheld."])
    ws2.append(["Methods X, Y and Z are three detection methods; which is which is not disclosed."])
    ws2.append([])
    ws2.append(["Outcome", "When to choose it"]); ws2["A6"].font = Font(bold=True)
    ws2["B6"].font = Font(bold=True)
    RB = {"IS_associated_supported": "a transposase ORF with at least one resolved terminal inverted repeat, or a complete element reported by a boundary-based method",
          "integron_associated_supported": "an integrase with at least one attC site, or a complete integron structure",
          "multiple_MGE_evidence_supported": "both IS and integron evidence present in the block",
          "chromosomal_mobile_supported": "credible mobile-element evidence that you cannot assign exclusively to IS or integron",
          "chromosomal_quiescent_supported": "no credible mobile-element evidence within the block",
          "neither_classification_supported": "evidence is present but contradicts both readings",
          "biologically_indeterminate": "the evidence shown cannot decide. USE THIS RATHER THAN GUESSING; it is scored as its own category, not as an error."}
    for o in OUTCOMES:
        ws2.append([o, RB[o]])
    ws2.append([])
    ws2.append(["boundary cases", "if an element appears truncated at a block edge, judge only "
                                  "what is visible and mark indeterminate if that is insufficient"])
    ws2.column_dimensions["A"].width = 40; ws2.column_dimensions["B"].width = 110
    for row in ws2.iter_rows():
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    xp = os.path.join(a.outdir, "NMV1_AUDIT_BLINDED_120.xlsx")
    wb.save(xp)

    hp = os.path.join(a.outdir, "NMV1_AUDIT_CASEBOOK.html")
    with open(hp, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("<!doctype html><meta charset='utf-8'><title>NM-V1 audit casebook</title>")
        fh.write("<style>body{font:14px/1.5 ui-sans-serif,system-ui;margin:24px;max-width:1060px}"
                 "h2{font-size:15px;margin:28px 0 4px;border-top:1px solid #ddd;padding-top:14px}"
                 "table{border-collapse:collapse;font-size:12px}td,th{border:1px solid #ddd;"
                 "padding:3px 7px}code{background:#f4f4f4;padding:1px 4px}"
                 ".warn{color:#8a4b00;font-weight:600}"
                 ".leg span{display:inline-block;width:12px;height:12px;vertical-align:-2px;"
                 "border:1px solid #444;margin-right:4px}</style>")
        fh.write("<h1>NM-V1 blinded expert-audit casebook</h1>")
        fh.write("<p>%d cases. Coordinates are 1-based, relative to the block. Methods X, Y and "
                 "Z are three detection methods; which is which is withheld, as are species, "
                 "study, gene identity, sampling stratum, audit stratum and every "
                 "machine-derived label.</p>" % len(cases))
        fh.write("<p class='leg'><b>Legend</b> "
                 "<span style='background:#9e9e9e'></span>resistance-gene interval "
                 "<span style='background:#7fb3d5'></span>transposase (X) "
                 "<span style='background:#c39bd3'></span>integrase (X) "
                 "<span style='background:#5d8aa8'></span>complete element (Y) "
                 "<span style='background:#a9cce3'></span>partial element (Y) "
                 "<span style='background:#f0b27a'></span>attC (Z) "
                 "<span style='background:#d5dbdb'></span>cassette protein (Z) "
                 "<span style='background:#1b2631'></span>terminal inverted repeat</p>")
        fh.write("<h2>Rubric</h2><table><tr><th>outcome</th><th>when</th></tr>")
        for o in OUTCOMES:
            fh.write("<tr><td><code>%s</code></td><td>%s</td></tr>" % (o, html.escape(RB[o])))
        fh.write("</table>")
        for t, b, L, av, F, g in cases:
            fh.write("<h2 id='%s'>%s</h2>" % (t, t))
            w = ""
            if g["truncated"] == "yes" or g["wrapped"] == "yes":
                w = " <span class='warn'>boundary warning: %s</span>" % (
                    "truncated" if g["truncated"] == "yes" else "wrapped across the origin")
            if g["tool_status"] != "ok":
                w += " <span class='warn'>tool status: %s</span>" % html.escape(g["tool_status"])
            fh.write("<p>length <code>%d bp</code> &middot; topology <code>%s</code> &middot; "
                     "gene intervals <code>%s</code>%s</p>"
                     % (L, g["topology"], "; ".join("%d-%d" % (x[0], x[1]) for x in av[:6]) or "none", w))
            fh.write(svg(L, av, F))
            fh.write("<table><tr><th>method</th><th>feature</th><th>label</th><th>start</th>"
                     "<th>end</th><th>strand</th><th>e-value</th><th>dist</th><th>detail</th></tr>")
            for f in F:
                fh.write("<tr><td>%s</td><td>%s</td><td>%s</td><td>%d</td><td>%d</td><td>%s</td>"
                         "<td>%s</td><td>%s</td><td>%s</td></tr>"
                         % (f["m"], html.escape(f["t"]), html.escape(str(f["l"])[:40]), f["b"],
                            f["e"], html.escape(str(f["s"])), html.escape(str(f["v"])[:12]),
                            f["d"], html.escape(str(f["x"])[:120])))
            fh.write("</table>")

    kp = os.path.join(a.outdir, "NMV1_AUDIT_UNBLINDING_KEY.tsv")
    with open(kp, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("token\tblock_id\taudit_stratum\trule_id\trule_based_label\thmm_is\thmm_integron\tspecies\tbioproject\n")
        for t, b, L, av, F, g in cases:
            st = next((s for s in order if b in sel[s]), "")
            fh.write("%s\t%s\t%s\t%s\t%s\t%d\t%d\t%s\t%s\n"
                     % (t, b, st, g["rule_id"], g["rule_based_label"], g["hmm_is"],
                        g["hmm_integron"], g["species"], g["bioproject"]))

    banned = set()
    for b in audit:
        g = idx[b]
        banned.update({g["species"], g["bioproject"], b, g["stratum"], g["rule_id"],
                       g["rule_based_label"]})
    banned.discard("")
    htxt = open(hp, encoding="utf-8").read()
    import zipfile
    zx = zipfile.ZipFile(xp)
    xtxt = " ".join(zx.read(n).decode("utf-8", "replace") for n in zx.namelist() if n.endswith(".xml"))
    # rubric legitimately contains outcome names; test identity fields only
    ident = {x for x in banned if not x.startswith(("chromosomal_", "IS_", "integron_",
                                                    "multiple_", "neither_", "biologically_"))
             and x not in ("A", "B", "C", "D", "E", "F1", "F2", "F3", "F4")}
    lh = sum(1 for x in ident if x and x in htxt)
    lx = sum(1 for x in ident if x and x in xtxt)
    print("\n=== BLINDING CHECK ===")
    print("  identity strings tested : %d" % len(ident))
    print("  leaks in casebook       : %d" % lh)
    print("  leaks in XLSX           : %d" % lx)

    rec = {"builder": VERSION,
           "run_utc": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
           "engine_sha256": ENGINE_SHA,
           "ground_truth": {"n_blocks": len(gt), "by_rule": dict(cnt), "by_label": dict(lab),
                            "evaluable": sum(1 for r in gt if r["evaluable_status"] == "evaluable"),
                            "indeterminate": sum(1 for r in gt if r["evaluable_status"] == "indeterminate"),
                            "table_sha256": sha256_file(gp)},
           "audit": {"strata": {s: {"pool": len(pools[s]), "target": target[s],
                                    "allocated": alloc[s], "selected": len(sel[s])} for s in order},
                     "total": len(audit), "cap": E["expert_audit"]["max_cases"],
                     "manifest_sha256": sha256_file(amp),
                     "xlsx_sha256": sha256_file(xp), "casebook_sha256": sha256_file(hp),
                     "unblinding_key_sha256": sha256_file(kp),
                     "key_sealed": True, "audit_performed": False},
           "blinding_check": {"strings_tested": len(ident), "leaks_html": lh, "leaks_xlsx": lx},
           "statements": ["The HMM label was not an input to any rule.",
                          "No aggregate performance metric was computed.",
                          "No unblinding key was opened.",
                          "No tool was rerun."]}
    rp = os.path.join(a.outdir, "NMV1_RULE_ENGINE_APPLICATION_RECEIPT.json")
    json.dump(rec, open(rp, "w", encoding="utf-8", newline="\n"), indent=2)
    print("\n  XLSX     : %s" % sha256_file(xp))
    print("  CASEBOOK : %s" % sha256_file(hp))
    print("  KEY      : %s" % sha256_file(kp))
    print("  RECEIPT  : %s" % sha256_file(rp))


if __name__ == "__main__":
    main()
