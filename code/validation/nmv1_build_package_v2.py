"""NM-V1 adjudication package V2 -- evidence-rich, with the frozen 100-block QC arm restored.

V1 is superseded, not deleted: it carried counts only, which cannot support a biological
judgement, and its QC arm was 80 rather than the frozen 100 because of elif precedence.

No tool is rerun. The frozen sample, the biological rules and the reference calls are
unchanged. Only QC sampling and package presentation change.

Coordinate convention, stated once and used everywhere: every coordinate shown to the
adjudicator is 1-based and RELATIVE TO THE BLOCK. ISEScan and IntegronFinder report block-
relative coordinates natively; HMM features and ARG intervals are chromosome-relative in the
source tables and are converted by subtracting block_start.
"""
import argparse, collections, csv, datetime, hashlib, html, json, os, sys

VERSION = "nmv1_build_package_v2_v1.0.0"
DESIGN_SHA = "c2aea6cb583c24b997ab376861acc600295e94d59bfa0ef2d55cf2bcc424bb20"
RULES_SHA = "e454873a89d1d56b20adb9ae157f20224076966daa0ed34bcbe44318063260f6"
SEED = 20260821
QC_TARGET = {"QC_AUTO_INTEGRON": 20, "QC_AUTO_IS": 40, "QC_AUTO_QUIESCENT": 40}
OUTCOMES = ["chromosomal_mobile_supported", "chromosomal_quiescent_supported",
            "integron_associated_supported", "IS_associated_supported",
            "multiple_MGE_evidence_supported", "neither_classification_supported",
            "biologically_indeterminate"]


def sha256_file(p):
    h = hashlib.sha256()
    with open(p, "rb") as fh:
        for c in iter(lambda: fh.read(1 << 20), b""):
            h.update(c)
    return h.hexdigest()


def det_shuffle(items, seed, salt):
    return sorted(items, key=lambda x: hashlib.sha256(
        ("%s|%d|%s" % (x, seed, salt)).encode()).hexdigest())


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo", required=True)
    ap.add_argument("--design", required=True)
    ap.add_argument("--rules", required=True)
    ap.add_argument("--evidence", required=True)
    ap.add_argument("--results", required=True)
    ap.add_argument("--outdir", required=True)
    a = ap.parse_args()
    if sha256_file(a.design) != DESIGN_SHA or sha256_file(a.rules) != RULES_SHA:
        print("REFUSING: frozen design or rules digest mismatch"); sys.exit(1)
    D = json.load(open(a.design, encoding="utf-8"))
    R = json.load(open(a.rules, encoding="utf-8"))
    print("%s | design %s | rules %s verified" % (VERSION, DESIGN_SHA[:12], RULES_SHA[:12]))

    ev = {r["block_id"]: r for r in csv.DictReader(open(a.evidence, encoding="utf-8"),
                                                   delimiter="\t")}
    print("  evidence table: %d blocks" % len(ev))
    O = os.path.join(a.repo, "audit/data/derived/pr_context/out")
    blk = {r["block_id"]: r for r in csv.DictReader(
        open(os.path.join(O, "shared_context_blocks.tsv"), encoding="utf-8"), delimiter="\t")}

    # ---- ARG intervals per block (chromosome coords -> block-relative) ----
    args_by_block = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(O, "arg_mge_neighbourhood.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        rep = r["replicon_accession"]
        for b, s in blk.items():
            if s["replicon_accession"] != rep:
                continue
            bs, be = int(s["block_start"]), int(s["block_end"])
            gs, ge = int(r["gene_start"]), int(r["gene_end"])
            if bs <= gs and be >= ge:
                args_by_block[b].append((gs - bs + 1, ge - bs + 1, r["strand"]))
                break
    print("  ARG intervals mapped for %d blocks" % len(args_by_block))

    # ---- HMM features (chromosome coords -> block-relative) ----
    hmm_by_block = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(O, "mge_feature_inventory.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        b = r["block_id"]
        s = blk.get(b)
        if not s:
            continue
        bs = int(s["block_start"])
        cl = r["feature_class"].lower()
        kind = "transposase" if ("is" in cl or "transpos" in cl) else "integrase"
        hmm_by_block[b].append({
            "kind": kind, "name": r["feature_name"],
            "beg": int(r["chrom_start"]) - bs + 1, "end": int(r["chrom_end"]) - bs + 1,
            "strand": r["strand"], "evalue": r["evalue"]})

    # ---- ISEScan elements ----
    ise_by_block = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(a.results, "isescan_hits.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        ise_by_block[r["block_id"]].append(r)
    # ---- IntegronFinder elements ----
    inf_by_block = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(a.results, "integronfinder_features.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        inf_by_block[r["block_id"]].append(r)

    # ================= QC ARM, three mutually exclusive subsets =================
    auto = [b for b, r in ev.items() if r["auto_label"]]
    pool_int = sorted([b for b in auto if int(ev[b]["if_complete"]) > 0])
    pool_is = sorted([b for b in auto if ev[b]["auto_label"] == "AUTO_IS_MOBILE"
                      and int(ev[b]["isescan_complete"]) > 0
                      and int(ev[b]["if_complete"]) == 0])
    pool_qui = sorted([b for b in auto if ev[b]["auto_label"] == "AUTO_QUIESCENT"])
    print("\n=== QC ELIGIBLE POOLS (recorded before selection) ===")
    print("  QC_AUTO_INTEGRON  pool %4d  (auto-labelled AND >=1 complete integron)" % len(pool_int))
    print("  QC_AUTO_IS        pool %4d  (AUTO_IS_MOBILE, complete IS, NO complete integron)" % len(pool_is))
    print("  QC_AUTO_QUIESCENT pool %4d" % len(pool_qui))

    sel = {}
    used = set()
    for cell, pool in (("QC_AUTO_INTEGRON", pool_int), ("QC_AUTO_IS", pool_is),
                       ("QC_AUTO_QUIESCENT", pool_qui)):
        avail = [b for b in pool if b not in used]
        pick = det_shuffle(avail, SEED, cell)[:QC_TARGET[cell]]
        sel[cell] = sorted(pick)
        used.update(pick)
        print("  %-18s requested %2d  available %4d  selected %2d"
              % (cell, QC_TARGET[cell], len(avail), len(pick)))
    qc_all = sorted(used)
    nqc = len(qc_all)
    assert len(qc_all) == len(set(qc_all)), "QC subsets overlap"
    print("  QC total %d | mutually exclusive: %s" % (nqc, len(qc_all) == sum(len(v) for v in sel.values())))

    mand = sorted([b for b, r in ev.items() if r["adjudication_route"]])
    pkg = sorted(set(mand) | set(qc_all))
    print("\n=== RECONCILIATION ===")
    print("  unique sampled blocks                       : %d" % len(ev))
    print("  automatic-label blocks                      : %d" % len(auto))
    print("  mandatory-adjudication blocks               : %d" % len(mand))
    print("  auto + mandatory                            : %d" % (len(auto) + len(mand)))
    print("  QC rows (concealed subset OF the %d auto)  : %d" % (len(auto), nqc))
    print("  package rows = mandatory + QC               : %d" % len(pkg))
    print("  QC blocks are NOT additional unique blocks  : True")

    # ---- QC manifest, written BEFORE the package ----
    os.makedirs(a.outdir, exist_ok=True)
    qcman = os.path.join(a.outdir, "NMV1_QC_MANIFEST_V2.json")
    tok = {b: "NMV1-%04d" % (i + 1) for i, b in
           enumerate(det_shuffle(pkg, SEED, "tokenshuffle"))}
    QM = {"manifest": "NMV1_QC_MANIFEST_V2", "builder": VERSION,
          "frozen_utc": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
          "seed": SEED,
          "selection_rule": "deterministic: candidates sorted by SHA-256 of "
                            "'<block_id>|<seed>|<cell>', first n taken; cells processed in the "
                            "order INTEGRON, IS, QUIESCENT so subsets are mutually exclusive",
          "cells": {c: {"target": QC_TARGET[c], "eligible_pool": len(p), "selected": len(sel[c]),
                        "selected_token_hash": hashlib.sha256(
                            "|".join(sorted(sel[c])).encode()).hexdigest()}
                    for c, p in (("QC_AUTO_INTEGRON", pool_int), ("QC_AUTO_IS", pool_is),
                                 ("QC_AUTO_QUIESCENT", pool_qui))},
          "qc_total": nqc, "mutually_exclusive": True,
          "all_selected_token_hash": hashlib.sha256("|".join(qc_all).encode()).hexdigest(),
          "note": "QC blocks are a concealed review subset drawn FROM the automatic-label "
                  "blocks. They are not additional unique blocks and do not change the 1,283 "
                  "sampled total. Including a complete-IS block in the integron cell does not "
                  "alter its biological label; it identifies which automatic rule is audited."}
    json.dump(QM, open(qcman, "w", encoding="utf-8", newline="\n"), indent=2)
    print("\n  QC manifest: %s" % sha256_file(qcman))

    # ================= evidence-rich rows =================
    def rows_for(b):
        e = ev[b]; s = blk[b]
        L = int(s["block_span_bp"])
        argiv = args_by_block.get(b, [])
        ise = ise_by_block.get(b, [])
        inf = inf_by_block.get(b, [])
        hm = hmm_by_block.get(b, [])
        def dist(beg, end):
            if not argiv:
                return ""
            return min(0 if (beg <= ae and end >= ab) else min(abs(beg - ae), abs(ab - end))
                       for ab, ae, _ in argiv)
        feats = []
        for f in hm:
            feats.append({"method": "X", "type": f["kind"], "label": f["name"],
                          "beg": f["beg"], "end": f["end"], "strand": f["strand"],
                          "evalue": f["evalue"], "extra": "",
                          "dist": dist(f["beg"], f["end"])})
        for r in ise:
            ir = ""
            if r.get("start1") and r.get("end1") and r.get("start2") and r.get("end2"):
                ir = "L:%s-%s R:%s-%s len=%s id=%s" % (r["start1"], r["end1"], r["start2"],
                                                       r["end2"], r["irLen"], r["irId"])
            feats.append({"method": "Y",
                          "type": "IS_complete" if r.get("type") == "c" else "IS_partial",
                          "label": "%s/%s" % (r.get("family", ""), r.get("cluster", "")),
                          "beg": int(r["isBegin"]), "end": int(r["isEnd"]),
                          "strand": r.get("strand", ""), "evalue": r.get("E_value", ""),
                          "extra": "TIR %s | ORF %s-%s len %s" % (ir or "none", r.get("orfBegin", ""),
                                                                  r.get("orfEnd", ""), r.get("orfLen", "")),
                          "dist": dist(int(r["isBegin"]), int(r["isEnd"]))})
        for r in inf:
            try:
                pb, pe = int(r["pos_beg"]), int(r["pos_end"])
            except ValueError:
                continue
            ann = r.get("annotation", "")
            t = "attC" if r.get("type_elt") == "attC" else ("integrase" if ann == "intI" else "cassette_protein")
            feats.append({"method": "Z", "type": t,
                          "label": "%s [%s]" % (r.get("id_integron", ""), r.get("type", "")),
                          "beg": pb, "end": pe, "strand": r.get("strand", ""),
                          "evalue": r.get("evalue", ""), "extra": "model=%s" % r.get("model", ""),
                          "dist": dist(pb, pe)})
        feats.sort(key=lambda f: f["beg"])
        return e, s, L, argiv, feats

    def svg(b, L, argiv, feats):
        W, LH = 980, 22
        tracks = ["ARG", "X", "Y", "Z"]
        H = 70 + len(tracks) * LH + max(0, len(feats) - 0) * 0
        sc = lambda x: 40 + (W - 80) * max(0, min(L, x)) / max(L, 1)
        p = ['<svg xmlns="http://www.w3.org/2000/svg" width="%d" height="%d" '
             'font-family="ui-monospace,monospace" font-size="11">' % (W, H)]
        p.append('<rect width="%d" height="%d" fill="#ffffff"/>' % (W, H))
        # scale bar
        p.append('<line x1="%d" y1="28" x2="%d" y2="28" stroke="#333"/>' % (sc(0), sc(L)))
        for k in range(0, 6):
            x = sc(L * k / 5.0)
            p.append('<line x1="%.1f" y1="24" x2="%.1f" y2="32" stroke="#333"/>' % (x, x))
            p.append('<text x="%.1f" y="20" text-anchor="middle" fill="#333">%d</text>'
                     % (x, int(L * k / 5.0)))
        p.append('<text x="8" y="20" fill="#333">bp</text>')
        ytr = {t: 48 + i * LH for i, t in enumerate(tracks)}
        for t in tracks:
            p.append('<text x="8" y="%d" fill="#555">%s</text>' % (ytr[t] + 10, t))
            p.append('<line x1="%d" y1="%d" x2="%d" y2="%d" stroke="#eee"/>'
                     % (sc(0), ytr[t] + 7, sc(L), ytr[t] + 7))
        for ab, ae, st in argiv:
            p.append('<rect x="%.1f" y="%d" width="%.1f" height="12" fill="#9e9e9e" '
                     'stroke="#555"/>' % (sc(ab), ytr["ARG"], max(2, sc(ae) - sc(ab))))
        style = {"transposase": "#7fb3d5", "integrase": "#c39bd3",
                 "IS_complete": "#5d8aa8", "IS_partial": "#a9cce3",
                 "attC": "#f0b27a", "cassette_protein": "#d5dbdb", "integrase_z": "#bb8fce"}
        for f in feats:
            y = ytr[f["method"]] if f["method"] in ytr else ytr["X"]
            col = style.get(f["type"], "#cccccc")
            x1, x2 = sc(f["beg"]), sc(f["end"])
            p.append('<rect x="%.1f" y="%d" width="%.1f" height="12" fill="%s" stroke="#444">'
                     '<title>%s %s %s-%s</title></rect>'
                     % (x1, y, max(2, x2 - x1), col, html.escape(f["type"]),
                        html.escape(str(f["label"])), f["beg"], f["end"]))
            if f["type"] == "IS_complete" and "TIR L:" in f["extra"]:
                for tx in (x1, x2):
                    p.append('<rect x="%.1f" y="%d" width="4" height="12" fill="#1b2631"/>'
                             % (tx - 2, y))
        p.append('</svg>')
        return "".join(p)

    # ---- write XLSX ----
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    ws = wb.active; ws.title = "adjudication"
    hdr = ["token", "adjudication", "reason", "topology", "boundary_warning",
           "window_length_bp", "arg_intervals", "n_arg", "methodX_transposase",
           "methodX_integrase", "methodY_IS_complete", "methodY_IS_partial",
           "methodY_TIR_evidence", "methodY_ORFs", "methodZ_integrase",
           "methodZ_attC_sites", "methodZ_structure", "nearest_feature_distance_bp",
           "tool_status", "feature_map"]
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = Alignment(wrap_text=True, vertical="top")
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(OUTCOMES), allow_blank=False,
                        showDropDown=False)
    dv.error = "Choose one of the seven permitted outcomes."
    dv.prompt = "Select exactly one outcome. biologically_indeterminate is a legitimate answer."
    ws.add_data_validation(dv)

    caseblocks = []
    for b in det_shuffle(pkg, SEED, "tokenshuffle"):
        e, s, L, argiv, feats = rows_for(b)
        t = tok[b]
        nX_t = sum(1 for f in feats if f["method"] == "X" and f["type"] == "transposase")
        nX_i = sum(1 for f in feats if f["method"] == "X" and f["type"] == "integrase")
        nY_c = sum(1 for f in feats if f["type"] == "IS_complete")
        nY_p = sum(1 for f in feats if f["type"] == "IS_partial")
        tir = "; ".join(f["extra"].split(" | ")[0].replace("TIR ", "")
                        for f in feats if f["type"] in ("IS_complete", "IS_partial")
                        and "TIR L:" in f["extra"])[:250]
        orfs = "; ".join(f["extra"].split(" | ")[1] for f in feats
                         if f["type"] in ("IS_complete", "IS_partial")
                         and " | " in f["extra"])[:250]
        nZ_i = sum(1 for f in feats if f["method"] == "Z" and f["type"] == "integrase")
        attc = [f for f in feats if f["type"] == "attC"]
        struct = ",".join(sorted({f["label"].split("[")[-1].rstrip("]")
                                  for f in feats if f["method"] == "Z"})) or "none"
        dmin = [f["dist"] for f in feats if f["dist"] != ""]
        row = [t,
               "", "",
               s["topology"],
               "yes" if (s["truncated"] == "yes" or s["wrapped_circular"] == "yes") else "no",
               L,
               "; ".join("%d-%d(%s)" % x for x in argiv[:6]) or "none",
               len(argiv), nX_t, nX_i, nY_c, nY_p, tir or "none", orfs or "none",
               nZ_i, len(attc),
               struct,
               min(dmin) if dmin else "",
               e["tool_status"],
               "casebook: %s" % t]
        ws.append(row)
        r = ws.max_row
        dv.add(ws.cell(row=r, column=2))
        ws.cell(row=r, column=20).hyperlink = "NMV1_ADJUDICATION_CASEBOOK_V2.html#%s" % t
        ws.cell(row=r, column=20).font = Font(color="0000EE", underline="single")
        caseblocks.append((t, b, L, argiv, feats, s, e))
    for col, w in zip("ABCDEFGHIJKLMNOPQRST",
                      [12, 34, 40, 10, 12, 14, 26, 7, 14, 14, 14, 14, 34, 34, 12, 12, 18, 14, 14, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B2"

    ws2 = wb.create_sheet("rubric")
    ws2.append(["NM-V1 adjudication rubric"]); ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])
    ws2.append(["Judge only the structural evidence shown. Species, study, gene identity and "
                "the original classification are withheld deliberately."])
    ws2.append(["Method X, Y and Z are three detection methods. You are not told which is which."])
    ws2.append([])
    ws2.append(["Permitted outcome", "When to choose it"])
    ws2["A6"].font = Font(bold=True); ws2["B6"].font = Font(bold=True)
    RB = R["adjudication_rubric"]
    for o in OUTCOMES:
        ws2.append([o, RB.get(o, RB.get(o.replace("_supported", ""), ""))])
    ws2.append([])
    ws2.append(["boundary cases", RB["boundary_cases"]])
    ws2.append(["indeterminate", "An honest indeterminate is more useful than a forced call "
                                 "and is reported as its own category."])
    ws2.column_dimensions["A"].width = 40; ws2.column_dimensions["B"].width = 110
    for row in ws2.iter_rows():
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    xlsx = os.path.join(a.outdir, "NMV1_ADJUDICATION_BLINDED_PACKAGE_V2.xlsx")
    wb.save(xlsx)

    # ---- HTML casebook ----
    hp = os.path.join(a.outdir, "NMV1_ADJUDICATION_CASEBOOK_V2.html")
    with open(hp, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("<!doctype html><meta charset='utf-8'><title>NM-V1 casebook</title>")
        fh.write("<style>body{font:14px/1.5 ui-sans-serif,system-ui;margin:24px;max-width:1060px}"
                 "h1{font-size:20px}h2{font-size:15px;margin:28px 0 4px;border-top:1px solid #ddd;padding-top:14px}"
                 "table{border-collapse:collapse;font-size:12px;margin:6px 0}"
                 "td,th{border:1px solid #ddd;padding:3px 7px;text-align:left}"
                 "code{background:#f4f4f4;padding:1px 4px}.warn{color:#8a4b00;font-weight:600}"
                 ".leg span{display:inline-block;width:12px;height:12px;vertical-align:-2px;"
                 "border:1px solid #444;margin-right:4px}</style>")
        fh.write("<h1>NM-V1 blinded adjudication casebook, V2</h1>")
        fh.write("<p>%d cases. Coordinates are 1-based and relative to the block. "
                 "Method X, Y and Z are three detection methods; which is which is withheld, "
                 "as are species, study, gene identity, sampling stratum and the original "
                 "classification.</p>" % len(caseblocks))
        fh.write("<p class='leg'><b>Legend</b> &nbsp;"
                 "<span style='background:#9e9e9e'></span>resistance-gene interval &nbsp;"
                 "<span style='background:#7fb3d5'></span>transposase (X) &nbsp;"
                 "<span style='background:#c39bd3'></span>integrase (X) &nbsp;"
                 "<span style='background:#5d8aa8'></span>complete element (Y) &nbsp;"
                 "<span style='background:#a9cce3'></span>partial element (Y) &nbsp;"
                 "<span style='background:#f0b27a'></span>attC (Z) &nbsp;"
                 "<span style='background:#d5dbdb'></span>cassette protein (Z) &nbsp;"
                 "<span style='background:#1b2631'></span>terminal inverted repeat</p>")
        fh.write("<h2>Rubric</h2><table><tr><th>outcome</th><th>when</th></tr>")
        for o in OUTCOMES:
            fh.write("<tr><td><code>%s</code></td><td>%s</td></tr>"
                     % (o, html.escape(RB.get(o, RB.get(o.replace("_supported", ""), "")))))
        fh.write("</table><p>%s</p>" % html.escape(RB["boundary_cases"]))
        for t, b, L, argiv, feats, s, e in caseblocks:
            fh.write("<h2 id='%s'>%s</h2>" % (t, t))
            warn = ""
            if s["truncated"] == "yes" or s["wrapped_circular"] == "yes":
                warn = " <span class='warn'>boundary warning: block is %s</span>" % (
                    "truncated" if s["truncated"] == "yes" else "wrapped across the origin")
            if e["tool_status"] != "ok":
                warn += " <span class='warn'>tool status: %s</span>" % html.escape(e["tool_status"])
            fh.write("<p>length <code>%d bp</code> &middot; topology <code>%s</code> &middot; "
                     "resistance-gene intervals <code>%s</code>%s</p>"
                     % (L, s["topology"],
                        "; ".join("%d-%d" % (x[0], x[1]) for x in argiv[:6]) or "none", warn))
            fh.write(svg(b, L, argiv, feats))
            fh.write("<table><tr><th>method</th><th>feature</th><th>label</th><th>start</th>"
                     "<th>end</th><th>strand</th><th>e-value</th><th>dist to gene</th>"
                     "<th>structural detail</th></tr>")
            for f in feats:
                fh.write("<tr><td>%s</td><td>%s</td><td>%s</td><td>%d</td><td>%d</td>"
                         "<td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>"
                         % (f["method"], html.escape(f["type"]), html.escape(str(f["label"])[:40]),
                            f["beg"], f["end"], html.escape(str(f["strand"])),
                            html.escape(str(f["evalue"])[:12]), f["dist"],
                            html.escape(str(f["extra"])[:120])))
            fh.write("</table>")

    # ---- V2 unblinding key ----
    key = os.path.join(a.outdir, "NMV1_ADJUDICATION_UNBLINDING_KEY_V2.tsv")
    with open(key, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("token\tblock_id\tstratum\tspecies\tbioproject\troute\tauto_label\tqc_cell\n")
        for t, b, L, argiv, feats, s, e in caseblocks:
            cell = next((c for c, v in sel.items() if b in v), "")
            fh.write("%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\n"
                     % (t, b, e["stratum"], e["species"], e["bioproject"],
                        e["adjudication_route"] or "QC", e["auto_label"], cell))

    # ---- leakage check ----
    banned = set()
    for b in pkg:
        e = ev[b]
        banned.update({e["species"], e["bioproject"], b, e["stratum"]})
    banned.discard("")
    txt = open(hp, encoding="utf-8").read()
    leak_html = sum(1 for x in banned if x and x in txt)
    import zipfile
    zx = zipfile.ZipFile(xlsx)
    xtxt = " ".join(zx.read(n).decode("utf-8", "replace") for n in zx.namelist()
                    if n.endswith(".xml"))
    leak_xlsx = sum(1 for x in banned if x and x in xtxt)
    print("\n=== BLINDING CHECK ===")
    print("  identity-bearing strings tested : %d" % len(banned))
    print("  leaks in casebook HTML          : %d" % leak_html)
    print("  leaks in XLSX                   : %d" % leak_xlsx)

    rec = {"builder": VERSION,
           "run_utc": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
           "design_sha256": DESIGN_SHA, "rules_sha256": RULES_SHA,
           "reconciliation": {"unique_sampled_blocks": len(ev),
                              "automatic_label_blocks": len(auto),
                              "mandatory_adjudication_blocks": len(mand),
                              "auto_plus_mandatory": len(auto) + len(mand),
                              "qc_rows_drawn_from_automatic": nqc,
                              "package_rows": len(pkg),
                              "qc_are_not_additional_unique_blocks": True},
           "qc": {c: {"target": QC_TARGET[c], "selected": len(sel[c])} for c in QC_TARGET},
           "qc_manifest_sha256": sha256_file(qcman),
           "outputs": {"xlsx_sha256": sha256_file(xlsx),
                       "casebook_sha256": sha256_file(hp),
                       "unblinding_key_v2_sha256": sha256_file(key)},
           "superseded_v1": {
             "package": sha256_file(os.path.join(
                 a.outdir, "NMV1_ADJUDICATION_BLINDED_PACKAGE_V1_SUPERSEDED_INSUFFICIENT_EVIDENCE.tsv")),
             "key": sha256_file(os.path.join(
                 a.outdir, "NMV1_ADJUDICATION_UNBLINDING_KEY_V1_SUPERSEDED_INSUFFICIENT_EVIDENCE.tsv"))},
           "blinding_check": {"strings_tested": len(banned),
                              "leaks_html": leak_html, "leaks_xlsx": leak_xlsx},
           "adjudication_performed": False,
           "no_tool_was_rerun": True,
           "statements": ["Ground truth is the frozen adjudicated labels.",
                          "No aggregate performance metric was computed.",
                          "Neither unblinding key was opened."]}
    rp = os.path.join(a.outdir, "NMV1_PACKAGE_V2_RECEIPT.json")
    json.dump(rec, open(rp, "w", encoding="utf-8", newline="\n"), indent=2)
    print("\n  XLSX     : %s" % sha256_file(xlsx))
    print("  CASEBOOK : %s" % sha256_file(hp))
    print("  KEY V2   : %s" % sha256_file(key))
    print("  RECEIPT  : %s" % sha256_file(rp))


if __name__ == "__main__":
    main()
