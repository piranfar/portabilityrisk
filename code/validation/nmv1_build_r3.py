"""NM-V1 R3 audit package.

Fixes two defects in R2: hyperlinks pointed at the R1 casebook, and the disclosed case
NMV1A-001 remained in the workbook leaving only 119 independently scored cases.

The replacement is fixed by NMV1_R3_REPLACEMENT_RULE.json, declared and hashed before any
evidence of the replacement was read. R3 uses a fresh token salt so no R3 token maps back to an
R2 token and the replacement is indistinguishable from any other case.

No tool is rerun. The frozen sample, rule engine and biological evidence are untouched.
"""
import argparse, collections, csv, datetime, hashlib, html, json, os, sys

VERSION = "nmv1_build_r3_v1.0.0"
ENGINE_SHA = "ed5db383bb0afe1a1a8433886d6666fe72c324975de99c6763a37824d51c2bee"
RULE_SHA = "49818469d56f75227b1c5548e39935a022da2a3d05d6f85d5a46276559122a8e"
TOKEN_SALT = "r3token"
TOKEN_PREFIX = "NMV1B"
OUTCOMES = ["chromosomal_mobile_supported", "chromosomal_quiescent_supported",
            "integron_associated_supported", "IS_associated_supported",
            "multiple_MGE_evidence_supported", "neither_classification_supported",
            "biologically_indeterminate"]
NEUTRAL = {
 "IS_associated_supported":
   "the block contains credible insertion-sequence evidence, for example a transposase ORF, "
   "terminal inverted repeats, or an element reported as complete by a boundary-based method. "
   "Whether what you see is sufficient is your judgement.",
 "integron_associated_supported":
   "credible integron evidence, for example an integrase, attC sites, or a complete integron "
   "structure. Sufficiency is your judgement.",
 "multiple_MGE_evidence_supported":
   "credible evidence of both insertion-sequence and integron elements in the same block.",
 "chromosomal_mobile_supported":
   "credible mobile-element evidence that you cannot assign exclusively to IS or integron.",
 "chromosomal_quiescent_supported": "no credible mobile-element evidence within the block.",
 "neither_classification_supported": "evidence is present but contradicts both readings.",
 "biologically_indeterminate":
   "evidence is present but you judge it too fragmentary, ambiguous or incomplete to establish "
   "or exclude mobile context, or the panel simply cannot decide. A legitimate answer, scored "
   "as its own category and never as an error."}
NO_THRESHOLD = (
 "No completeness threshold is prescribed. Some elements are reported as complete and some as "
 "partial; some have one terminal inverted repeat resolved and some both. Where you draw the "
 "line between sufficient and insufficient evidence is precisely what this audit measures, so "
 "record your genuine judgement rather than trying to match any rule.")
BOUNDARY = ("if an element appears truncated at a block edge, judge only what is visible and "
            "choose biologically_indeterminate if the visible part is insufficient.")
SCORING = ("Primary agreement is scored on three states: MOBILE (the four mobile outcomes), "
           "QUIESCENT, and NON-EVALUABLE. Choosing a more general mobile outcome instead of a "
           "specific one does not count against agreement.")


def sha(p):
    h = hashlib.sha256()
    with open(p, "rb") as fh:
        for c in iter(lambda: fh.read(1 << 20), b""):
            h.update(c)
    return h.hexdigest()


def dsort(x, seed, salt):
    return sorted(x, key=lambda b: hashlib.sha256(
        ("%s|%d|%s" % (b, seed, salt)).encode()).hexdigest())


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo", required=True)
    ap.add_argument("--dir", required=True)
    a = ap.parse_args()
    D = a.dir
    if sha(os.path.join(D, "NMV1_RULE_ENGINE_FROZEN.json")) != ENGINE_SHA:
        print("REFUSING: engine digest mismatch"); sys.exit(1)
    if sha(os.path.join(D, "NMV1_R3_REPLACEMENT_RULE.json")) != RULE_SHA:
        print("REFUSING: replacement rule digest mismatch"); sys.exit(1)
    E = json.load(open(os.path.join(D, "NMV1_RULE_ENGINE_FROZEN.json"), encoding="utf-8"))
    M = json.load(open(os.path.join(D, "NMV1_AUDIT_MANIFEST.json"), encoding="utf-8"))
    RR = json.load(open(os.path.join(D, "NMV1_R3_REPLACEMENT_RULE.json"), encoding="utf-8"))
    SEED = E["expert_audit"]["seed"]
    print("%s | engine + replacement rule verified" % VERSION)

    gt = {r["block_id"]: r for r in csv.DictReader(
        open(os.path.join(D, "NMV1_RULE_BASED_GROUND_TRUTH.tsv"), encoding="utf-8"),
        delimiter="\t")}
    ev = {r["block_id"]: r for r in csv.DictReader(
        open(os.path.join(D, "nmv1_block_evidence_table.tsv"), encoding="utf-8"),
        delimiter="\t")}

    def n(x):
        try:
            return int(x)
        except (TypeError, ValueError):
            return 0

    MOB = {"A", "B", "C", "D"}
    pools = {
      "S_IS": [b for b, r in gt.items() if r["rule_id"] == "B"],
      "S_INT": [b for b, r in gt.items() if r["rule_id"] == "C"],
      "S_MULTI": [b for b, r in gt.items() if r["rule_id"] == "A"],
      "S_QUIET": [b for b, r in gt.items() if r["rule_id"] == "E"],
      "S_DISC": [b for b, r in gt.items() if r["rule_id"] in MOB | {"E"}
                 and ((n(ev[b]["hmm_is"]) + n(ev[b]["hmm_integron"])) > 0) != (r["rule_id"] in MOB)],
      "S_INDET": [b for b, r in gt.items() if r["rule_id"].startswith("F")]}
    order = [s["id"] for s in E["expert_audit"]["strata"]]
    sel = {}; used = set()
    for s in order:
        avail = [b for b in sorted(pools[s]) if b not in used]
        pick = dsort(avail, SEED, s)[:M["strata"][s]["allocated"]]
        sel[s] = sorted(pick); used.update(pick)
    audit_r2 = sorted(used)
    if hashlib.sha256("|".join(audit_r2).encode()).hexdigest() != M["all_token_hash"]:
        print("REFUSING: could not reproduce the R2 selection"); sys.exit(2)
    old_tok = {b: "NMV1A-%03d" % (i + 1) for i, b in enumerate(dsort(audit_r2, SEED, "audittoken"))}
    disclosed = [b for b, t in old_tok.items() if t == "NMV1A-001"][0]

    ordered = dsort(sorted(pools["S_INDET"]), SEED, "S_INDET")
    repl = ordered[M["strata"]["S_INDET"]["allocated"]]
    if hashlib.sha256(disclosed.encode()).hexdigest() != RR["replaced_block_sha256"] \
       or hashlib.sha256(repl.encode()).hexdigest() != RR["replacement_block_sha256"]:
        print("REFUSING: replacement does not match the hashed declaration"); sys.exit(3)
    print("  replaced/replacement block hashes match the declaration")
    assert repl not in audit_r2, "replacement was already in the audit set"
    audit = sorted((set(audit_r2) - {disclosed}) | {repl})
    assert len(audit) == 120 and disclosed not in audit
    print("  R3 scored set: %d blocks | disclosed case removed | replacement added" % len(audit))

    tok = {b: "%s-%03d" % (TOKEN_PREFIX, i + 1)
           for i, b in enumerate(dsort(audit, SEED, TOKEN_SALT))}

    O = os.path.join(a.repo, "audit/data/derived/pr_context/out")
    RES = os.path.join(a.repo, "audit/data/derived/nmv1")
    blk = {r["block_id"]: r for r in csv.DictReader(
        open(os.path.join(O, "shared_context_blocks.tsv"), encoding="utf-8"), delimiter="\t")}
    ise = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(RES, "isescan_hits.tsv"), encoding="utf-8"),
                            delimiter="\t"):
        ise[r["block_id"]].append(r)
    inf = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(RES, "integronfinder_features.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        inf[r["block_id"]].append(r)
    argb = collections.defaultdict(list)
    S = set(audit)
    for r in csv.DictReader(open(os.path.join(O, "arg_mge_neighbourhood.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        for b in S:
            s = blk[b]
            if s["replicon_accession"] != r["replicon_accession"]:
                continue
            bs, be = int(s["block_start"]), int(s["block_end"])
            gs, ge = int(r["gene_start"]), int(r["gene_end"])
            if bs <= gs and be >= ge:
                argb[b].append((gs - bs + 1, ge - bs + 1, r["strand"]))
                break
    hmmb = collections.defaultdict(list)
    for r in csv.DictReader(open(os.path.join(O, "mge_feature_inventory.tsv"),
                                 encoding="utf-8"), delimiter="\t"):
        b = r["block_id"]
        if b not in S:
            continue
        bs = int(blk[b]["block_start"])
        cl = r["feature_class"].lower()
        hmmb[b].append({"kind": "transposase" if ("is" in cl or "transpos" in cl) else "integrase",
                        "name": r["feature_name"], "beg": int(r["chrom_start"]) - bs + 1,
                        "end": int(r["chrom_end"]) - bs + 1, "strand": r["strand"],
                        "evalue": r["evalue"]})

    def feats(b):
        L = int(blk[b]["block_span_bp"]); av = argb.get(b, [])
        def d(x, y):
            if not av:
                return ""
            return min(0 if (x <= ae and y >= ab) else min(abs(x - ae), abs(ab - y))
                       for ab, ae, _ in av)
        F = []
        for f in hmmb.get(b, []):
            F.append({"m": "X", "t": f["kind"], "l": f["name"], "b": f["beg"], "e": f["end"],
                      "s": f["strand"], "v": f["evalue"], "x": "", "d": d(f["beg"], f["end"])})
        for h in ise.get(b, []):
            ir = ("L:%s-%s R:%s-%s len=%s id=%s" % (h["start1"], h["end1"], h["start2"],
                                                    h["end2"], h["irLen"], h["irId"])
                  if n(h.get("irLen")) > 0 else "none")
            F.append({"m": "Y", "t": "IS_complete" if h.get("type") == "c" else "IS_partial",
                      "l": "%s/%s" % (h.get("family", ""), h.get("cluster", "")),
                      "b": n(h["isBegin"]), "e": n(h["isEnd"]), "s": h.get("strand", ""),
                      "v": h.get("E_value", ""),
                      "x": "TIR %s | ORF %s-%s len %s" % (ir, h.get("orfBegin", ""),
                                                          h.get("orfEnd", ""), h.get("orfLen", "")),
                      "d": d(n(h["isBegin"]), n(h["isEnd"]))})
        for f in inf.get(b, []):
            pb, pe = n(f["pos_beg"]), n(f["pos_end"])
            if not pb and not pe:
                continue
            t = "attC" if f.get("type_elt") == "attC" else (
                "integrase" if f.get("annotation") == "intI" else "cassette_protein")
            F.append({"m": "Z", "t": t, "l": "%s [%s]" % (f.get("id_integron", ""), f.get("type", "")),
                      "b": pb, "e": pe, "s": f.get("strand", ""), "v": f.get("evalue", ""),
                      "x": "model=%s" % f.get("model", ""), "d": d(pb, pe)})
        F.sort(key=lambda z: z["b"])
        return L, av, F

    def svg(L, av, F):
        W, LH = 980, 22
        tr = ["ARG", "X", "Y", "Z"]; H = 60 + len(tr) * LH
        sc = lambda x: 40 + (W - 80) * max(0, min(L, x)) / max(L, 1)
        p = ['<svg xmlns="http://www.w3.org/2000/svg" width="%d" height="%d" '
             'font-family="ui-monospace,monospace" font-size="11">' % (W, H),
             '<rect width="%d" height="%d" fill="#fff"/>' % (W, H),
             '<line x1="%.1f" y1="28" x2="%.1f" y2="28" stroke="#333"/>' % (sc(0), sc(L))]
        for k in range(6):
            x = sc(L * k / 5.0)
            p.append('<line x1="%.1f" y1="24" x2="%.1f" y2="32" stroke="#333"/>' % (x, x))
            p.append('<text x="%.1f" y="20" text-anchor="middle" fill="#333">%d</text>'
                     % (x, int(L * k / 5.0)))
        p.append('<text x="8" y="20" fill="#333">bp</text>')
        yt = {t: 44 + i * LH for i, t in enumerate(tr)}
        for t in tr:
            p.append('<text x="8" y="%d" fill="#555">%s</text>' % (yt[t] + 10, t))
            p.append('<line x1="%.1f" y1="%d" x2="%.1f" y2="%d" stroke="#eee"/>'
                     % (sc(0), yt[t] + 7, sc(L), yt[t] + 7))
        for ab, ae, _ in av:
            p.append('<rect x="%.1f" y="%d" width="%.1f" height="12" fill="#9e9e9e" stroke="#555"/>'
                     % (sc(ab), yt["ARG"], max(2, sc(ae) - sc(ab))))
        col = {"transposase": "#7fb3d5", "integrase": "#c39bd3", "IS_complete": "#5d8aa8",
               "IS_partial": "#a9cce3", "attC": "#f0b27a", "cassette_protein": "#d5dbdb"}
        for f in F:
            y = yt.get(f["m"], yt["X"]); x1, x2 = sc(f["b"]), sc(f["e"])
            p.append('<rect x="%.1f" y="%d" width="%.1f" height="12" fill="%s" stroke="#444">'
                     '<title>%s %s-%s</title></rect>'
                     % (x1, y, max(2, x2 - x1), col.get(f["t"], "#ccc"),
                        html.escape(f["t"]), f["b"], f["e"]))
            if f["t"] == "IS_complete" and "TIR L:" in f["x"]:
                for tx in (x1, x2):
                    p.append('<rect x="%.1f" y="%d" width="4" height="12" fill="#1b2631"/>' % (tx - 2, y))
        p.append('</svg>')
        return "".join(p)

    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, Alignment
    CASEBOOK = "NMV1_AUDIT_CASEBOOK_R3.html"
    wb = Workbook(); ws = wb.active; ws.title = "audit"
    hdr = ["token", "adjudication", "reason", "topology", "boundary_warning", "window_length_bp",
           "arg_intervals", "methodX_transposase", "methodX_integrase", "methodY_IS_complete",
           "methodY_IS_partial", "methodY_bilateral_TIR", "methodY_TIR_detail", "methodY_ORFs",
           "methodZ_integrase", "methodZ_attC", "methodZ_structure",
           "nearest_feature_distance_bp", "tool_status", "feature_map"]
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = Alignment(wrap_text=True, vertical="top")
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(OUTCOMES), allow_blank=False)
    dv.prompt = "Exactly one outcome. biologically_indeterminate is legitimate."
    ws.add_data_validation(dv)
    cases = []
    for b in dsort(audit, SEED, TOKEN_SALT):
        L, av, F = feats(b); g = ev[b]; t = tok[b]
        bil = sum(1 for f in F if f["t"] == "IS_complete" and "TIR L:" in f["x"])
        tir = "; ".join(f["x"].split(" | ")[0][4:] for f in F
                        if f["t"].startswith("IS") and "TIR L:" in f["x"])[:250]
        orfs = "; ".join(f["x"].split(" | ")[1] for f in F
                         if f["t"].startswith("IS") and " | " in f["x"])[:250]
        struct = ",".join(sorted({f["l"].split("[")[-1].rstrip("]")
                                  for f in F if f["m"] == "Z"})) or "none"
        ds = [f["d"] for f in F if f["d"] != ""]
        ws.append([t, "", "", g["topology"],
                   "yes" if (g["truncated"] == "yes" or g["wrapped"] == "yes") else "no", L,
                   "; ".join("%d-%d" % (x[0], x[1]) for x in av[:6]) or "none",
                   sum(1 for f in F if f["m"] == "X" and f["t"] == "transposase"),
                   sum(1 for f in F if f["m"] == "X" and f["t"] == "integrase"),
                   sum(1 for f in F if f["t"] == "IS_complete"),
                   sum(1 for f in F if f["t"] == "IS_partial"), bil,
                   tir or "none", orfs or "none",
                   sum(1 for f in F if f["m"] == "Z" and f["t"] == "integrase"),
                   sum(1 for f in F if f["t"] == "attC"), struct,
                   min(ds) if ds else "", g["tool_status"], "casebook: %s" % t])
        r = ws.max_row
        dv.add(ws.cell(row=r, column=2))
        c = ws.cell(row=r, column=20)
        c.hyperlink = "%s#%s" % (CASEBOOK, t)
        c.font = Font(color="0000EE", underline="single")
        cases.append((t, b, L, av, F, g))
    for col, w in zip("ABCDEFGHIJKLMNOPQRST",
                      [12, 34, 44, 10, 12, 14, 24, 13, 13, 13, 13, 13, 32, 32, 12, 10, 16, 14, 14, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B2"
    ws2 = wb.create_sheet("rubric")
    ws2.append(["NM-V1 expert-audit rubric (R3)"]); ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])
    ws2.append(["Judge only the structural evidence shown. Species, study, gene identity, the "
                "original classification and every machine-derived label are withheld."])
    ws2.append(["Methods X, Y and Z are three detection methods; which is which is not disclosed."])
    ws2.append([]); ws2.append([NO_THRESHOLD]); ws2["A6"].font = Font(bold=True)
    ws2.append([]); ws2.append(["Outcome", "When to choose it"])
    ws2["A8"].font = Font(bold=True); ws2["B8"].font = Font(bold=True)
    for o in OUTCOMES:
        ws2.append([o, NEUTRAL[o]])
    ws2.append([]); ws2.append(["boundary cases", BOUNDARY]); ws2.append(["scoring", SCORING])
    ws2.column_dimensions["A"].width = 40; ws2.column_dimensions["B"].width = 112
    for row in ws2.iter_rows():
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    xp = os.path.join(D, "NMV1_AUDIT_BLINDED_120_R3.xlsx"); wb.save(xp)

    hp = os.path.join(D, CASEBOOK)
    with open(hp, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("<!doctype html><meta charset='utf-8'><title>NM-V1 audit casebook R3</title>")
        fh.write("<style>body{font:14px/1.5 ui-sans-serif,system-ui;margin:24px;max-width:1060px}"
                 "h2{font-size:15px;margin:28px 0 4px;border-top:1px solid #ddd;padding-top:14px}"
                 "table{border-collapse:collapse;font-size:12px}td,th{border:1px solid #ddd;"
                 "padding:3px 7px}code{background:#f4f4f4;padding:1px 4px}"
                 ".warn{color:#8a4b00;font-weight:600}.leg span{display:inline-block;width:12px;"
                 "height:12px;vertical-align:-2px;border:1px solid #444;margin-right:4px}</style>")
        fh.write("<h1>NM-V1 blinded expert-audit casebook (R3)</h1>")
        fh.write("<p>%d cases. Coordinates are 1-based, relative to the block. Methods X, Y and "
                 "Z are three detection methods; which is which is withheld, as are species, "
                 "study, gene identity and every machine-derived label.</p>" % len(cases))
        fh.write("<p class='leg'><b>Legend</b> <span style='background:#9e9e9e'></span>"
                 "resistance-gene interval <span style='background:#7fb3d5'></span>transposase (X) "
                 "<span style='background:#c39bd3'></span>integrase (X) "
                 "<span style='background:#5d8aa8'></span>complete element (Y) "
                 "<span style='background:#a9cce3'></span>partial element (Y) "
                 "<span style='background:#f0b27a'></span>attC (Z) "
                 "<span style='background:#d5dbdb'></span>cassette protein (Z) "
                 "<span style='background:#1b2631'></span>terminal inverted repeat</p>")
        fh.write("<h2>Rubric</h2><p><b>%s</b></p><table><tr><th>outcome</th><th>when</th></tr>"
                 % html.escape(NO_THRESHOLD))
        for o in OUTCOMES:
            fh.write("<tr><td><code>%s</code></td><td>%s</td></tr>" % (o, html.escape(NEUTRAL[o])))
        fh.write("</table><p>%s</p><p>%s</p>" % (html.escape(BOUNDARY), html.escape(SCORING)))
        for t, b, L, av, F, g in cases:
            fh.write("<h2 id='%s'>%s</h2>" % (t, t))
            w = ""
            if g["truncated"] == "yes" or g["wrapped"] == "yes":
                w = " <span class='warn'>boundary warning: %s</span>" % (
                    "truncated" if g["truncated"] == "yes" else "wrapped across the origin")
            if g["tool_status"] != "ok":
                w += " <span class='warn'>tool status: %s</span>" % html.escape(g["tool_status"])
            fh.write("<p>length <code>%d bp</code> &middot; topology <code>%s</code> &middot; "
                     "gene intervals <code>%s</code>%s</p>"
                     % (L, g["topology"], "; ".join("%d-%d" % (x[0], x[1]) for x in av[:6]) or "none", w))
            fh.write(svg(L, av, F))
            fh.write("<table><tr><th>method</th><th>feature</th><th>label</th><th>start</th>"
                     "<th>end</th><th>strand</th><th>e-value</th><th>dist</th><th>detail</th></tr>")
            for f in F:
                fh.write("<tr><td>%s</td><td>%s</td><td>%s</td><td>%d</td><td>%d</td><td>%s</td>"
                         "<td>%s</td><td>%s</td><td>%s</td></tr>"
                         % (f["m"], html.escape(f["t"]), html.escape(str(f["l"])[:40]), f["b"],
                            f["e"], html.escape(str(f["s"])), html.escape(str(f["v"])[:12]),
                            f["d"], html.escape(str(f["x"])[:120])))
            fh.write("</table>")

    kp = os.path.join(D, "NMV1_AUDIT_UNBLINDING_KEY_R3.tsv")
    with open(kp, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("token\tblock_id\taudit_stratum\trule_id\trule_based_label\tis_replacement\n")
        for t, b, L, av, F, g in cases:
            st = next((s for s in order if b in sel[s]), "S_INDET")
            fh.write("%s\t%s\t%s\t%s\t%s\t%s\n"
                     % (t, b, st, gt[b]["rule_id"], gt[b]["rule_based_label"],
                        "yes" if b == repl else "no"))
    dp = os.path.join(D, "NMV1_DISCLOSED_EXCLUDED.tsv")
    with open(dp, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("r2_token\tblock_sha256\taudit_stratum\treason\tin_r3_scoring_set\n")
        fh.write("NMV1A-001\t%s\tS_INDET\tmachine label discussed with the adjudicator before "
                 "review; blinding lost\tno\n" % hashlib.sha256(disclosed.encode()).hexdigest())

    man = os.path.join(D, "NMV1_AUDIT_MANIFEST_R3.json")
    MM = {"manifest": "NMV1_AUDIT_MANIFEST_R3", "builder": VERSION,
          "frozen_utc": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
          "engine_sha256": ENGINE_SHA, "replacement_rule_sha256": RULE_SHA,
          "token_prefix": TOKEN_PREFIX, "token_salt": TOKEN_SALT, "seed": SEED,
          "scored_cases": len(audit),
          "disclosed_excluded": {"r2_token": "NMV1A-001",
                                 "block_sha256": RR["replaced_block_sha256"],
                                 "in_scoring_set": False},
          "replacement": {"block_sha256": RR["replacement_block_sha256"],
                          "stratum": "S_INDET", "selected_by": "hashed deterministic rule",
                          "evidence_consulted_before_selection": False},
          "strata": {s: len([b for b in audit if b in sel[s]]) for s in order},
          "all_token_hash": hashlib.sha256("|".join(sorted(audit)).encode()).hexdigest(),
          "note": "S_INDET count includes the replacement, which is not in sel['S_INDET']"}
    MM["strata"]["S_INDET"] = len([b for b in audit if b in sel["S_INDET"] or b == repl])
    json.dump(MM, open(man, "w", encoding="utf-8", newline="\n"), indent=2)

    # ---------------- verification ----------------
    print("\n=== VERIFICATION ===")
    import openpyxl, zipfile, re
    w2 = openpyxl.load_workbook(xp); s2 = w2["audit"]
    toks = [s2.cell(row=r, column=1).value for r in range(2, s2.max_row + 1)]
    htxt = open(hp, encoding="utf-8").read()
    anchors = set(re.findall(r"<h2 id='(%s-\d{3})'>" % TOKEN_PREFIX, htxt))
    ok = True

    def C(name, cond, detail=""):
        nonlocal ok
        ok &= bool(cond)
        print("  %-52s %s %s" % (name, "PASS" if cond else "*** FAIL ***", detail))
    C("exactly 120 unique scored tokens", len(toks) == 120 and len(set(toks)) == 120, len(toks))
    C("zero disclosed cases among the 120",
      all(t.startswith(TOKEN_PREFIX) for t in toks) and "NMV1A-001" not in toks)
    C("all adjudication cells blank",
      sum(1 for r in range(2, s2.max_row + 1) if s2.cell(row=r, column=2).value) == 0)
    C("all reason cells blank",
      sum(1 for r in range(2, s2.max_row + 1) if s2.cell(row=r, column=3).value) == 0)
    C("dropdown validation on all 120 cells",
      len(str(s2.data_validations.dataValidation[0].sqref).split()) == 120)
    tg = [s2.cell(row=r, column=20).hyperlink.target for r in range(2, s2.max_row + 1)
          if s2.cell(row=r, column=20).hyperlink]
    C("every hyperlink targets the R3 filename",
      len(tg) == 120 and all(t.split("#")[0] == CASEBOOK for t in tg))
    C("every hyperlink anchor exists in the casebook",
      all(t.split("#")[1] in anchors for t in tg))
    C("Excel and HTML token sets identical", set(toks) == anchors, "%d vs %d" % (len(set(toks)), len(anchors)))
    C("120 svg feature maps", htxt.count("<svg") == 120)
    banned = set()
    for b in audit:
        banned.update({ev[b]["species"], ev[b]["bioproject"], b, ev[b]["stratum"],
                       gt[b]["rule_id"], gt[b]["rule_based_label"]})
    ident = {x for x in banned if x and not x.startswith(
        ("chromosomal_", "IS_", "integron_", "multiple_", "neither_", "biologically_"))
        and x not in {"A", "B", "C", "D", "E", "F1", "F2", "F3", "F4"}}
    zx = zipfile.ZipFile(xp)
    xt = " ".join(zx.read(nm).decode("utf-8", "replace") for nm in zx.namelist() if nm.endswith(".xml"))
    lh = [x for x in ident if x in htxt]; lx = [x for x in ident if x in xt]
    C("zero identity leakage in casebook", not lh, "%d tested" % len(ident))
    C("zero identity leakage in xlsx", not lx)
    for f, h in (("NMV1_AUDIT_BLINDED_120.xlsx", "f66dfea818dc5002cdeab64321ba5db28bcba6fcad7cc453cead79228f2759a4"),
                 ("NMV1_AUDIT_CASEBOOK.html", "4e5531f6b37434757459bae46701e5c5ac5d429a7694d6b051b9838410299e55"),
                 ("NMV1_AUDIT_BLINDED_120_R2.xlsx", "dd4d32edb023c792493c93598a4263dce8ded3a3d4f1b901cdceb7b02c5c4bee"),
                 ("NMV1_AUDIT_CASEBOOK_R2.html", "6eef49f7fea5cd19f1d9847b58883d045a4e3db74a16af36300c2259862d6b47"),
                 ("NMV1_AUDIT_UNBLINDING_KEY.tsv", "b7042b95365a939132f9c190093a6a871326095ef0cb629b90e8a6ef77639929"),
                 ("NMV1_ADJUDICATION_UNBLINDING_KEY_V2.tsv", "54409bc5f704b2de29bca15199b2a1fc718ac888aea4640dcf7472c5f7cb459e"),
                 ("NMV1_RULE_ENGINE_FROZEN.json", ENGINE_SHA),
                 ("NMV1_RULE_BASED_GROUND_TRUTH.tsv", "1beecaa39048f4df52a3235f2dbc538056af9adc6912ee057cbac8ca55b85897")):
        C("preserved unchanged: %s" % f, sha(os.path.join(D, f)) == h)

    rec = {"builder": VERSION,
           "run_utc": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
           "engine_sha256": ENGINE_SHA, "replacement_rule_sha256": RULE_SHA,
           "scored_cases": 120, "disclosed_excluded": "NMV1A-001",
           "outputs": {"xlsx": sha(xp), "casebook": sha(hp), "key_r3": sha(kp),
                       "manifest_r3": sha(man), "disclosed_excluded_tsv": sha(dp)},
           "verification_all_passed": bool(ok),
           "no_tool_rerun": True, "no_key_opened": True,
           "no_performance_or_agreement_computed": True}
    rp = os.path.join(D, "NMV1_R3_RECEIPT.json")
    json.dump(rec, open(rp, "w", encoding="utf-8", newline="\n"), indent=2)
    print("\n  R3 XLSX     : %s" % sha(xp))
    print("  R3 CASEBOOK : %s" % sha(hp))
    print("  R3 KEY      : %s  (sealed)" % sha(kp))
    print("  R3 MANIFEST : %s" % sha(man))
    print("  DISCLOSED   : %s" % sha(dp))
    print("  R3 RECEIPT  : %s" % sha(rp))
    print("\n  ALL VERIFICATION: %s" % ("PASS" if ok else "*** FAIL ***"))
    sys.exit(0 if ok else 8)


if __name__ == "__main__":
    main()
