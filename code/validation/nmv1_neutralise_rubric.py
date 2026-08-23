"""NM-V1 Amendment 004 -- rubric neutralisation.

Rewrites ONLY the rubric wording shown to the adjudicator. Every token, every evidence value,
every feature map and the case ordering are copied through byte-for-byte from the existing
package. The originals are preserved; new R2 files are written alongside them.

The neutral rubric deliberately prescribes no completeness threshold. Where the adjudicator
draws the line between sufficient and insufficient evidence is the quantity this audit
measures; stating the engine's cutoff would make the audit circular.
"""
import argparse, copy, hashlib, html, json, os, re, sys

VERSION = "nmv1_neutralise_rubric_v1.0.0"
XLSX_SHA = "f66dfea818dc5002cdeab64321ba5db28bcba6fcad7cc453cead79228f2759a4"
HTML_SHA = "4e5531f6b37434757459bae46701e5c5ac5d429a7694d6b051b9838410299e55"

OUTCOMES = ["chromosomal_mobile_supported", "chromosomal_quiescent_supported",
            "integron_associated_supported", "IS_associated_supported",
            "multiple_MGE_evidence_supported", "neither_classification_supported",
            "biologically_indeterminate"]

NEUTRAL = {
 "IS_associated_supported":
   "the block contains credible insertion-sequence evidence, for example a transposase ORF, "
   "terminal inverted repeats, or an element reported as complete by a boundary-based method. "
   "Whether what you see is sufficient is your judgement.",
 "integron_associated_supported":
   "credible integron evidence, for example an integrase, attC sites, or a complete integron "
   "structure. Sufficiency is your judgement.",
 "multiple_MGE_evidence_supported":
   "credible evidence of both insertion-sequence and integron elements in the same block.",
 "chromosomal_mobile_supported":
   "credible mobile-element evidence that you cannot assign exclusively to IS or integron.",
 "chromosomal_quiescent_supported":
   "no credible mobile-element evidence within the block.",
 "neither_classification_supported":
   "evidence is present but contradicts both readings.",
 "biologically_indeterminate":
   "evidence is present but you judge it too fragmentary, ambiguous or incomplete to establish "
   "or exclude mobile context, or the panel simply cannot decide. A legitimate answer, scored "
   "as its own category and never as an error.",
}

NO_THRESHOLD = (
 "No completeness threshold is prescribed. Some elements are reported as complete and some as "
 "partial; some have one terminal inverted repeat resolved and some both. Where you draw the "
 "line between sufficient and insufficient evidence is precisely what this audit measures, so "
 "record your genuine judgement rather than trying to match any rule.")

BOUNDARY = ("if an element appears truncated at a block edge, judge only what is visible and "
            "choose biologically_indeterminate if the visible part is insufficient.")

EXCLUDED_TOKEN = "NMV1A-001"
EXCLUSION_NOTE = ("This case was discussed with the adjudicator before review and its machine "
                  "label is therefore known. Still answer it, but it is excluded from the "
                  "primary agreement denominator.")


def sha256_file(p):
    h = hashlib.sha256()
    with open(p, "rb") as fh:
        for c in iter(lambda: fh.read(1 << 20), b""):
            h.update(c)
    return h.hexdigest()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True)
    a = ap.parse_args()
    D = a.dir
    xin = os.path.join(D, "NMV1_AUDIT_BLINDED_120.xlsx")
    hin = os.path.join(D, "NMV1_AUDIT_CASEBOOK.html")
    if sha256_file(xin) != XLSX_SHA or sha256_file(hin) != HTML_SHA:
        print("REFUSING: source package digest mismatch"); sys.exit(1)
    xout = os.path.join(D, "NMV1_AUDIT_BLINDED_120_R2.xlsx")
    hout = os.path.join(D, "NMV1_AUDIT_CASEBOOK_R2.html")
    for p in (xout, hout):
        if os.path.exists(p):
            print("REFUSING: %s exists" % p); sys.exit(1)
    print("%s | source package verified" % VERSION)

    # ---------------- XLSX ----------------
    import openpyxl
    from openpyxl.styles import Font, Alignment
    wb = openpyxl.load_workbook(xin)
    ws = wb["audit"]
    before = [[c.value for c in r] for r in ws.iter_rows(min_row=1)]

    # add an exclusion note column, leaving every existing column untouched
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value="scoring_note").font = Font(bold=True)
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 46
    nex = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == EXCLUDED_TOKEN:
            ws.cell(row=r, column=col, value=EXCLUSION_NOTE).alignment = Alignment(wrap_text=True)
            nex += 1
    print("  exclusion note applied to %d row(s)" % nex)

    # rebuild the rubric sheet with neutral wording
    del wb["rubric"]
    ws2 = wb.create_sheet("rubric")
    ws2.append(["NM-V1 expert-audit rubric (revision 2)"])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])
    ws2.append(["Judge only the structural evidence shown. Species, study, gene identity, the "
                "original classification and every machine-derived label are withheld."])
    ws2.append(["Methods X, Y and Z are three detection methods; which is which is not disclosed."])
    ws2.append([])
    ws2.append([NO_THRESHOLD])
    ws2["A6"].font = Font(bold=True)
    ws2.append([])
    ws2.append(["Outcome", "When to choose it"])
    ws2["A8"].font = Font(bold=True); ws2["B8"].font = Font(bold=True)
    for o in OUTCOMES:
        ws2.append([o, NEUTRAL[o]])
    ws2.append([])
    ws2.append(["boundary cases", BOUNDARY])
    ws2.append(["scoring", "Primary agreement is scored on three states: MOBILE (the four "
                           "mobile outcomes), QUIESCENT, and NON-EVALUABLE (neither / "
                           "indeterminate). Choosing a more general mobile outcome instead of a "
                           "specific one does not count against agreement."])
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 112
    for row in ws2.iter_rows():
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(xout)

    # verify case data unchanged
    wb2 = openpyxl.load_workbook(xout)
    after = [[c.value for c in r[:20]] for r in wb2["audit"].iter_rows(min_row=1)]
    same = all(before[i][:20] == after[i] for i in range(len(before)))
    print("  case rows/columns 1-20 identical to source: %s" % same)
    print("  dropdown validations preserved: %d"
          % len(str(wb2["audit"].data_validations.dataValidation[0].sqref).split()))

    # ---------------- HTML ----------------
    t = open(hin, encoding="utf-8").read()
    old = re.search(r"<h2>Rubric</h2>.*?(?=<h2 id=)", t, re.S)
    if not old:
        print("REFUSING: could not locate the rubric block"); sys.exit(2)
    new = ["<h2>Rubric</h2>",
           "<p><b>%s</b></p>" % html.escape(NO_THRESHOLD),
           "<table><tr><th>outcome</th><th>when</th></tr>"]
    for o in OUTCOMES:
        new.append("<tr><td><code>%s</code></td><td>%s</td></tr>" % (o, html.escape(NEUTRAL[o])))
    new.append("</table>")
    new.append("<p>%s</p>" % html.escape(BOUNDARY))
    new.append("<p>Primary agreement is scored on three states: MOBILE (the four mobile "
               "outcomes), QUIESCENT, and NON-EVALUABLE. Choosing a more general mobile "
               "outcome instead of a specific one does not count against agreement.</p>")
    t2 = t[:old.start()] + "".join(new) + t[old.end():]
    # flag the excluded case in its own panel
    t2 = t2.replace("<h2 id='%s'>%s</h2>" % (EXCLUDED_TOKEN, EXCLUDED_TOKEN),
                    "<h2 id='%s'>%s</h2><p class='warn'>%s</p>"
                    % (EXCLUDED_TOKEN, EXCLUDED_TOKEN, html.escape(EXCLUSION_NOTE)))
    t2 = t2.replace("<h1>NM-V1 blinded expert-audit casebook</h1>",
                    "<h1>NM-V1 blinded expert-audit casebook (revision 2)</h1>")
    open(hout, "w", encoding="utf-8", newline="\n").write(t2)

    anchors = len(re.findall(r"<h2 id=.NMV1A-\d{3}.>", t2))
    print("  casebook panels preserved: %d | svg panels: %d" % (anchors, t2.count("<svg")))
    print("  exclusion note in casebook: %d" % t2.count(EXCLUSION_NOTE[:40]))

    # leakage re-check: no engine threshold language leaked into the instrument
    banned = ["type = c", "type=c", "IS_strong", "bilateral TIRs and", "rule B", "rule F3",
              "F3", "IS_bilateral_TIR_n", "rule_id", "rule-based label"]
    lh = [b for b in banned if b in t2]
    zx = __import__("zipfile").ZipFile(xout)
    xt = " ".join(zx.read(n).decode("utf-8", "replace") for n in zx.namelist() if n.endswith(".xml"))
    lx = [b for b in banned if b in xt]
    print("  engine-threshold terms leaked into casebook: %s" % (lh or "none"))
    print("  engine-threshold terms leaked into xlsx    : %s" % (lx or "none"))

    print("\n  R2 XLSX     : %s" % sha256_file(xout))
    print("  R2 CASEBOOK : %s" % sha256_file(hout))
    print("  source XLSX     unchanged: %s" % (sha256_file(xin) == XLSX_SHA))
    print("  source CASEBOOK unchanged: %s" % (sha256_file(hin) == HTML_SHA))


if __name__ == "__main__":
    main()
