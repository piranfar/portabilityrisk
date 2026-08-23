"""NM-V1 audit scoring. Runs only after the returned adjudication has been frozen and hashed.

Applies NMV1_AUDIT_SCORING_SPECIFICATION.json exactly: three-state architecture as the primary
metric with a Wilson interval, exact seven-label agreement as a secondary subtype analysis, and
the integron cases as a descriptive Clopper-Pearson estimate that carries no gate.
"""
import argparse, collections, csv, datetime, hashlib, json, math, os, sys

VERSION = "nmv1_score_audit_v1.0.0"
SPEC_SHA = "55ab34bd7d1d6c66b4f13c56eedff72e5eb07e53b0ba744a2d176c4b4bdda704"
KEY_SHA = "5fd81dd8919b759d77df233287e58c71d4f59afaa9a690a09e6a7134f6145cb6"
ADJ_SHA = "479839f34dc994c12432e208006337256ed2fe16668dd7cfbb7893d38489e3a5"


def sha(p):
    h = hashlib.sha256()
    with open(p, "rb") as fh:
        for c in iter(lambda: fh.read(1 << 20), b""):
            h.update(c)
    return h.hexdigest()


def wilson(k, n, z=1.96):
    if n == 0:
        return (float("nan"),) * 3
    p = k / n
    d = 1 + z * z / n
    c = (p + z * z / (2 * n)) / d
    h = z / d * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n))
    return p, max(0.0, c - h), min(1.0, c + h)


def clopper(k, n):
    """Exact binomial interval without scipy: bisection on the binomial tail."""
    if n == 0:
        return (float("nan"),) * 3
    from math import comb
    def cdf(p, k, n):
        return sum(comb(n, i) * p ** i * (1 - p) ** (n - i) for i in range(k + 1))
    def solve(f, lo, hi):
        for _ in range(200):
            m = (lo + hi) / 2
            if f(m) > 0:
                lo = m
            else:
                hi = m
        return (lo + hi) / 2
    p = k / n
    low = 0.0 if k == 0 else solve(lambda x: cdf(x, k - 1, n) - 0.975, 0.0, 1.0)
    high = 1.0 if k == n else solve(lambda x: cdf(x, k, n) - 0.025, 0.0, 1.0)
    return p, low, high


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True)
    a = ap.parse_args()
    D = a.dir
    spec_p = os.path.join(D, "NMV1_AUDIT_SCORING_SPECIFICATION.json")
    key_p = os.path.join(D, "NMV1_AUDIT_UNBLINDING_KEY_R3.tsv")
    adj_p = os.path.join(D, "NMV1_AUDIT_ADJUDICATED_120_R3_RECEIVED.xlsx")
    for p, h, n in ((spec_p, SPEC_SHA, "scoring spec"), (key_p, KEY_SHA, "R3 key"),
                    (adj_p, ADJ_SHA, "frozen adjudication")):
        if sha(p) != h:
            print("REFUSING: %s digest mismatch" % n); sys.exit(1)
    S = json.load(open(spec_p, encoding="utf-8"))
    print("%s | spec, key and frozen adjudication all verified" % VERSION)
    print("  the adjudication was frozen and hashed BEFORE this key was opened")

    STATE = {}
    for st, outs in S["state_map"].items():
        for o in outs:
            STATE[o] = st

    import openpyxl
    ws = openpyxl.load_workbook(adj_p)["audit"]
    adj = {ws.cell(row=r, column=1).value: str(ws.cell(row=r, column=2).value).strip()
           for r in range(2, ws.max_row + 1)}
    ev = {ws.cell(row=r, column=1).value: {
            "isc": ws.cell(row=r, column=10).value or 0,
            "isp": ws.cell(row=r, column=11).value or 0,
            "bil": ws.cell(row=r, column=12).value or 0,
            "mx_t": ws.cell(row=r, column=8).value or 0,
            "mx_i": ws.cell(row=r, column=9).value or 0,
            "zi": ws.cell(row=r, column=15).value or 0,
            "za": ws.cell(row=r, column=16).value or 0}
          for r in range(2, ws.max_row + 1)}
    key = {r["token"]: r for r in csv.DictReader(open(key_p, encoding="utf-8"), delimiter="\t")}
    toks = sorted(adj)
    assert len(toks) == 120 and set(toks) == set(key)

    rows = []
    for t in toks:
        k = key[t]
        h_out = adj[t]; m_out = k["rule_based_label"]
        rows.append({"token": t, "human": h_out, "machine": m_out,
                     "h_state": STATE[h_out], "m_state": STATE[m_out],
                     "rule_id": k["rule_id"], "stratum": k["audit_stratum"],
                     "exact": h_out == m_out,
                     "state_match": STATE[h_out] == STATE[m_out], **ev[t]})

    # ---------------- PRIMARY ----------------
    n = len(rows); k = sum(1 for r in rows if r["state_match"])
    p, lo, hi = wilson(k, n)
    print("\n" + "=" * 74)
    print("PRIMARY - three-state architecture agreement (frozen metric)")
    print("=" * 74)
    print("  agreement : %d / %d = %.4f" % (k, n, p))
    print("  Wilson 95%% CI : [%.4f, %.4f]" % (lo, hi))
    verdict = ("SUCCESS" if (p >= 0.90 and lo >= 0.80) else
               "REVISE" if (0.80 <= p < 0.90 or 0.65 <= lo < 0.80) else "FAIL")
    print("  thresholds: SUCCESS p>=0.90 and lo>=0.80 | REVISE p 0.80-0.899 or lo 0.65-0.799")
    print("  PRIMARY VERDICT: %s" % verdict)

    print("\n  three-state confusion (rows = machine, cols = human)")
    states = ["MOBILE", "QUIESCENT", "NON_EVALUABLE"]
    print("  %-16s %10s %10s %14s %8s" % ("", *states, "total"))
    for ms in states:
        r_ = [x for x in rows if x["m_state"] == ms]
        print("  %-16s %10d %10d %14d %8d"
              % (ms, *[sum(1 for x in r_ if x["h_state"] == hs) for hs in states], len(r_)))
    print("  %-16s %10d %10d %14d %8d"
          % ("total", *[sum(1 for x in rows if x["h_state"] == hs) for hs in states], len(rows)))

    print("\n  state-specific agreement")
    ss = {}
    for st in states:
        sub = [x for x in rows if x["m_state"] == st]
        if not sub:
            continue
        kk = sum(1 for x in sub if x["state_match"])
        pp, l2, h2 = wilson(kk, len(sub))
        ss[st] = {"n": len(sub), "k": kk, "agreement": pp, "ci": [l2, h2]}
        print("     %-16s %3d/%3d = %.4f  [%.4f, %.4f]" % (st, kk, len(sub), pp, l2, h2))

    # ---------------- SECONDARY ----------------
    ke = sum(1 for r in rows if r["exact"])
    pe, le, he = wilson(ke, n)
    print("\n" + "=" * 74)
    print("SECONDARY - exact seven-label agreement (subtype resolution, NO gate)")
    print("=" * 74)
    print("  agreement : %d / %d = %.4f  Wilson [%.4f, %.4f]" % (ke, n, pe, le, he))
    sub_only = sum(1 for r in rows if r["state_match"] and not r["exact"])
    print("  same state but different subtype: %d cases" % sub_only)

    # ---------------- INTEGRON, descriptive ----------------
    ic = [r for r in rows if r["rule_id"] == "C"]
    ik = sum(1 for r in ic if r["state_match"])
    ip, il, ih = clopper(ik, len(ic))
    print("\n" + "=" * 74)
    print("INTEGRON cases - descriptive only, carries NO gate")
    print("=" * 74)
    print("  n = %d | state agreement %d/%d = %.4f" % (len(ic), ik, len(ic), ip))
    print("  exact Clopper-Pearson 95%% CI : [%.4f, %.4f]  (width %.3f)" % (il, ih, ih - il))
    print("  this interval cannot resolve integron-specific performance and does not")
    print("  determine the NM-V1 verdict")

    # ---------------- prespecified partial-element analysis ----------------
    print("\n" + "=" * 74)
    print("PRESPECIFIED partial-element evidentiary-threshold analysis (Amendment 004)")
    print("=" * 74)
    pe_cases = [r for r in rows if r["isp"] > 0 and r["isc"] == 0
                and r["zi"] == 0 and r["za"] == 0]
    pm = sum(1 for r in pe_cases if r["h_state"] == "MOBILE")
    print("  partial-only IS, no integron evidence : %d cases" % len(pe_cases))
    print("     adjudicator called MOBILE          : %d" % pm)
    print("     adjudicator called NON-EVALUABLE   : %d"
          % sum(1 for r in pe_cases if r["h_state"] == "NON_EVALUABLE"))
    print("     machine label was NON-EVALUABLE    : %d"
          % sum(1 for r in pe_cases if r["m_state"] == "NON_EVALUABLE"))

    # ---------------- where the disagreement actually is ----------------
    print("\n" + "=" * 74)
    print("DISAGREEMENT ANATOMY")
    print("=" * 74)
    dis = [r for r in rows if not r["state_match"]]
    print("  total state disagreements: %d" % len(dis))
    by = collections.Counter((r["m_state"], r["h_state"]) for r in dis)
    for (m, h), v in by.most_common():
        print("     machine %-14s -> human %-14s %3d" % (m, h, v))
    print("\n  disagreements by machine rule id:")
    for rid, v in collections.Counter(r["rule_id"] for r in dis).most_common():
        tot = sum(1 for r in rows if r["rule_id"] == rid)
        print("     rule %-3s %3d of %3d (%.0f%%)" % (rid, v, tot, 100 * v / tot))
    mx = [r for r in dis if r["m_state"] == "QUIESCENT" and r["h_state"] == "MOBILE"]
    withx = sum(1 for r in mx if (r["mx_t"] or 0) + (r["mx_i"] or 0) > 0)
    print("\n  machine QUIESCENT -> human MOBILE : %d cases" % len(mx))
    print("     of those, Method X showed >=1 marker : %d (%.0f%%)"
          % (withx, 100 * withx / len(mx) if mx else 0))

    rec = {"builder": VERSION,
           "run_utc": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
           "spec_sha256": SPEC_SHA, "key_sha256": KEY_SHA,
           "adjudication_sha256": ADJ_SHA,
           "adjudication_frozen_before_key_opened": True,
           "primary": {"metric": "three-state architecture agreement", "n": n, "k": k,
                       "agreement": p, "wilson_ci": [lo, hi], "verdict": verdict},
           "state_specific": ss,
           "secondary_exact_seven_label": {"n": n, "k": ke, "agreement": pe,
                                           "wilson_ci": [le, he],
                                           "same_state_different_subtype": sub_only,
                                           "carries_gate": False},
           "integron_descriptive": {"n": len(ic), "k": ik, "agreement": ip,
                                    "clopper_pearson_ci": [il, ih], "carries_gate": False},
           "partial_element_analysis": {"n": len(pe_cases), "human_mobile": pm},
           "disagreement": {"total": len(dis),
                            "by_transition": {"%s->%s" % t: v for t, v in by.items()},
                            "by_rule": dict(collections.Counter(r["rule_id"] for r in dis)),
                            "quiescent_to_mobile": len(mx),
                            "quiescent_to_mobile_with_methodX_marker": withx}}
    rp = os.path.join(D, "NMV1_AUDIT_SCORING_RECEIPT.json")
    json.dump(rec, open(rp, "w", encoding="utf-8", newline="\n"), indent=2)
    dp = os.path.join(D, "nmv1_audit_scored_cases.tsv")
    cols = ["token", "stratum", "rule_id", "machine", "human", "m_state", "h_state",
            "state_match", "exact", "isc", "isp", "bil", "mx_t", "mx_i", "zi", "za"]
    with open(dp, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\t".join(cols) + "\n")
        for r in sorted(rows, key=lambda x: x["token"]):
            fh.write("\t".join(str(r[c]) for c in cols) + "\n")
    print("\n  scored cases : %s" % sha(dp))
    print("  receipt      : %s" % sha(rp))


if __name__ == "__main__":
    main()
